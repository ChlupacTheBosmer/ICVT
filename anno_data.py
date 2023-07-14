# This file contains the annotation data file classes
#
# Example usage:
#
# # PRODUCT
# excel = annotation_watcher_file("excel/GR2_L2_LavSto2_fil.xlsx", True, True, True, True)
# # Take visit duration and the assembled timestamp
# filtered_data = excel.dataframe.loc[:, ['duration', 'ts']].values.tolist()
# annotation_data_array = filtered_data
#
#
# # PRODUCT
# csv = annotation_custom_file("excel/croplog_tst.xlsx")
# # convert to list
# filtered_data = csv.dataframe.loc[:, ['duration', 'ts']].values.tolist()
# annotation_data_array = filtered_data

import utils
import pandas as pd
import os
import openpyxl
import xlwings as xw
from typing import Dict, Callable

class Ancestor_annotation_file():

    def __init__(self, filepath):

        # Define logger
        self.logger = utils.log_define()

        # Define variables
        self.filepath = filepath

    def evaluate_string_formula(self, cell):

        # If the cell contains a number, return the value as is
        if isinstance(cell, (int, float)):
            return cell
        # If the cell contains an Excel formula, use openpyxl to evaluate it
        elif cell.startswith('='):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'].value = cell
            value = ws['A1'].value
            wb.close()
            return value
        # If the cell contains text, return the text as is
        else:
            return cell

    def convert_months(self, cell):
        cell = self.evaluate_string_formula(cell)
        months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                  'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        try:
            cell = months.get(cell)  # Get corresponding value from dictionary
        except:
            self.logger.warning("Invalid month abreviation in Excel file")
            cell = 1
        return cell

    def convert_year(self, cell):
        cell = int(self.evaluate_string_formula(cell))
        return cell

    def convert_time_data(self, cell):
        cell = int(self.evaluate_string_formula(cell))
        cell = "{:02d}".format(cell)  # Format the number with leading zeros
        return cell

    def convert_bool(self, cell):
        cell = self.evaluate_string_formula(cell)
        if isinstance(cell, int) and cell == 1:
            return True
        else:
            return False

    def construct_timestamp(self, dataframe):
        dataframe = dataframe.copy()
        dataframe.loc[:, 'ts'] = dataframe.iloc[:, 0:6].apply(lambda x: f"{x[0]}{x[1]}{x[2]}_{x[3]}_{x[4]}_{x[5]}", axis=1)
        return dataframe

    def save_temp_file(self, dataframe):

        # Save temp file
        temp_excel_dir = os.path.join("resources", "exc")
        utils.create_dir(temp_excel_dir)
        dataframe.to_excel(os.path.join(temp_excel_dir, "output_filtered_crop.xlsx"), index=False)

    def load_excel_file(self, cols, converters):
        # Read the Excel file, skipping the first two rows
        try:
            dataframe = pd.read_excel(self.filepath, usecols=cols, skiprows=2, header=None,
                                      converters=converters)
        except ValueError as e:
            self.logger.error(f"Error reading Excel file {self.filepath}. Error message: {e}")

            # Open the Excel workbook using xlwings
            workbook = xw.Book(self.filepath)
            sheet = workbook.sheets[0]

            # Remove any filters
            if sheet.api.AutoFilterMode:
                sheet.api.AutoFilterMode = False

            # Save to temporary file
            utils.create_dir(os.path.join("resources", "exc"))
            temp_filepath = os.path.join("resources", "exc", "temp.xlsx")
            workbook.save(temp_filepath)
            workbook.close()

            # Read with pandas
            try:
                dataframe = pd.read_excel(temp_filepath, usecols=cols, skiprows=2, header=None,
                                          converters=converters)
            except ValueError as e:
                self.logger.error(
                    f"Attempted to fix errors in Excel file {self.filepath}. Attempt failed. Error message: {e}. Please fix the errors manually and try again.")
                return None
            self.logger.info(
                f"Attempted to remove filters from Excel file {self.filepath}. Saved a copy of the file to {temp_filepath}")

        self.logger.debug(f"Retrieved dataframe from Excel:\n{dataframe}")

        return dataframe

class Annotation_watcher_file(Ancestor_annotation_file):
    def __init__(self, filepath, load_time_data: bool = True, load_visit_data: bool = True, load_visitor_data: bool = True, load_behavior_data: bool = True):

        # Define logger
        self.logger = utils.log_define()

        # Define variables
        self.filepath = filepath
        self.load_time_data = load_time_data
        self.load_visit_data = load_visit_data
        self.load_visitor_data = load_visitor_data
        self.load_behavior_data = load_behavior_data

        # Load the Excel file into dataframe
        self.dataframe = self.construct_dataframe()

    def construct_dataframe(self):

        # The default values of cols to be extracted are:
        # 0 - A - Year
        # 1 - B - Month - !
        # ...
        # 5 - F - Seconds
        # 15 - P - Visitor arrival - filter column
        # 18 - S - Visit duration in seconds
        # 19 - T - Time of departure - Hours
        # 20 - U - Time of departure - Minutes
        # 21 - V - Time of departure - Seconds
        # 23 - X - Insect species
        # 24 - Y - Insect Order

        # For more info see the example excel table in resources/exc there I wrote down the numbers of columns, datatypes etc.

        # Define logger
        self.logger.debug(f"Running function construct_dataframe()")

        # Define the number of cols for each data extraction module
        time_cols: list[int] = [0, 1, 2, 3, 4, 5]
        visit_cols: list[int] = [15, 18, 19, 20, 21]
        visitor_cols: list[int] = [23, 24]
        behavior_cols: list[int] = [27, 34, 36, 40, 46, 47, 48, 49]

        # Define convertors for each data extraction module
        time_converters: Dict[int, Callable] = {0: self.convert_year,
                                                1: self.convert_months,
                                                2: self.convert_time_data,
                                                3: self.convert_time_data,
                                                4: self.convert_time_data,
                                                5: self.convert_time_data}

        visit_converters: Dict[int, Callable] = {15: self.convert_bool,
                                                 18: self.evaluate_string_formula,
                                                 19: self.convert_time_data,
                                                 20: self.convert_time_data,
                                                 21: self.convert_time_data}

        visitor_converters: Dict[int, Callable] = {23: self.evaluate_string_formula,
                                                   24: self.evaluate_string_formula}

        behavior_converters: Dict[int, Callable] = {27: self.evaluate_string_formula,
                                                    34: self.convert_bool,
                                                    36: self.convert_bool,
                                                    40: self.convert_bool,
                                                    46: self.convert_bool,
                                                    47: self.evaluate_string_formula,
                                                    48: self.convert_bool,
                                                    49: self.evaluate_string_formula}

        modules = [(self.load_time_data, time_cols, time_converters),
                   (self.load_visit_data, visit_cols, visit_converters),
                   (self.load_visitor_data, visitor_cols, visitor_converters),
                   (self.load_behavior_data, behavior_cols, behavior_converters)]

        # Assemble the cols and converters to be used in reading the Excel file
        cols = []
        converters = {}
        for i, (module_bool, module_cols, module_converters) in enumerate(modules):
            if module_bool:
                cols += module_cols
                converters.update(module_converters)

        # Open the excel, resolve any issues and load datarame
        dataframe = self.load_excel_file(cols, converters)

        # Mapping dictionary for column renaming
        column_mapping = {
            0: 'year',
            1: 'month',
            2: 'day',
            3: 'hour_a',
            4: 'min_a',
            5: 'sec_a',
            15: 'visit',
            18: 'duration',
            19: 'hour_d',
            20: 'min_d',
            21: 'sec_d',
            23: 'vis_id',
            24: 'vis_ord',
            27: 'no_flo',
            34: 'f_pol',
            36: 'f_nec',
            40: 'f_fp',
            46: 'con_m',
            47: 'no_con_m',
            48: 'con_f',
            49: 'no_con_f'
        }

        # Rename columns using the mapping dictionary
        dataframe.rename(columns=column_mapping, inplace=True)

        # Filter data frame based on whether the value in the column of index 6 (P - visitor arrival) is 1.
        col_to_filter_by: str = "visit"
        filtered_dataframe = dataframe[dataframe[col_to_filter_by] == True]

        # Add another column called "ts" for timestamp
        filtered_dataframe = self.construct_timestamp(filtered_dataframe)

        # Debug check
        #print(f"Flow: Filtered dataframe:\n {filtered_dataframe}")

        # Preprocess the visitor descriptions if applicable
        if self.load_visitor_data:
            # Get the column names
            visitor_id = 'vis_id'
            visitor_ord = 'vis_ord'

            # Count the number of NAs in each column
            na_count_id = filtered_dataframe[visitor_id].isna().sum()
            na_count_ord = filtered_dataframe[visitor_ord].isna().sum()

            # Check which column has fewer NAs
            if na_count_id <= na_count_ord:
                chosen_column = visitor_id
                other_column = visitor_ord
            else:
                chosen_column = visitor_ord
                other_column = visitor_id

            # Replace NAs in chosen_column with values from other_column if they are not NAs
            filtered_dataframe = filtered_dataframe.copy()
            filtered_dataframe.loc[:, chosen_column].fillna(filtered_dataframe[other_column], inplace=True)
            filtered_dataframe.loc[:, other_column] = filtered_dataframe.loc[:, chosen_column]

        # Save temporary file from dataframe to excel
        self.save_temp_file(filtered_dataframe)

        # Return the dataframe
        return filtered_dataframe


class Annotation_custom_file(Ancestor_annotation_file):

    def __init__(self, filepath):

        # Define logger
        self.logger = utils.log_define()

        # Define variables
        self.filepath = filepath

        # Load the Excel file into dataframe
        self.dataframe = self.construct_dataframe()

    def construct_dataframe(self):

        # Define logger
        self.logger.debug(f"Running function construct_dataframe()")

        # Define the columns to extract
        cols: list[int] = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        converters: Dict[int, Callable] = {0: self.convert_year,
                                           1: self.convert_months,
                                           2: self.convert_time_data,
                                           3: self.convert_time_data,
                                           4: self.convert_time_data,
                                           5: self.convert_time_data,
                                           6: self.evaluate_string_formula,
                                           7: self.convert_time_data,
                                           8: self.convert_time_data,
                                           9: self.convert_time_data
                                           }

        # Read the Excel file, skipping the first two rows - follow the custom format
        dataframe = self.load_excel_file(cols, converters)

        self.logger.debug(f"Retrieved dataframe from Excel:\n{dataframe}")

        # Make a copy to solve the error when trying to access only a slice
        dataframe = dataframe.copy()

        # Mapping dictionary for column renaming
        column_mapping = {
            0: 'year',
            1: 'month',
            2: 'day',
            3: 'hour_a',
            4: 'min_a',
            5: 'sec_a',
            6: 'duration',
            7: 'hour_d',
            8: 'min_d',
            9: 'sec_d'
        }

        # Rename columns using the mapping dictionary
        dataframe.rename(columns=column_mapping, inplace=True)

        # Add another column called "ts" for timestamp
        dataframe = self.construct_timestamp(dataframe)

        # Save temporary file from dataframe to excel
        self.save_temp_file(dataframe)

        # Return the dataframe
        return dataframe


