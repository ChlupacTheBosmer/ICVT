# This file contains the ICCS app class that inherits from ICVT AppAncestor class
import utils
import icvt
import pandas as pd
import os
import subprocess
import re
import cv2
import pytesseract
import configparser
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageTk
import pickle
import datetime
import sys
import random
import openpyxl
import math
import time
from hachoir.metadata import extractMetadata
from hachoir.parser import createParser
import logging
import xlwings as xw
import keyboard
from ultralytics import YOLO
import torch
import shutil
import asyncio
import cProfile
import tracemalloc
from typing import Dict, Callable
import numpy as np

class ICCS(icvt.AppAncestor):
    def __init__(self):
        # Define logger
        self.logger = self.log_define()

        # Init basic instance variables and get config
        self.app_title = "Insect Communities Crop Suite"
        self.config = self.config_create()
        self.config_read()
        self.scanned_folders = []
        self.dir_hierarchy = False
        self.loaded = False
        self.auto = 0
        self.auto_processing = 0
        self.frames = []
        self.modified_frames = []
        self.video_filepaths = []
        self.points_of_interest_entry = []
        self.button_images = []
        self.buttons = []
        self.gui_imgs = []
        self.cap = None

        # Initiation functions - get directories and files
        self.scan_default_folders()
        while not self.check_path():
            self.get_video_folder(1)
        self.get_excel_path(1, 1)
        self.load_videos()
        self.reload_points_of_interest()
        self.load_video_frames()

        # Create output folders
        utils.create_dir(self.output_folder)
        utils.create_dir(f"./{self.output_folder}/whole frames/")

    def config_create(self):
        # Set default values
        config = configparser.ConfigParser()
        config['Resource Paths'] = {
            'OCR_tesseract_path': 'C:/Program Files/Tesseract-OCR/tesseract.exe',
            'video_folder_path': '',
            'annotation_file_path': '',
            'output_folder': 'output'
        }
        config['OCR settings'] = {
            'x_coordinate': '0',
            'y_coordinate': '0',
            'width': '500',
            'height': '40'
        }
        config['Crop settings'] = {
            'crop_mode': '1',
            'crop_interval_frames': '30',
            'frames_per_visit': '0',
            'filter_visitors': '0',
            'yolo_processing': '0',
            'default_label_category': '6',
            'yolo_conf': '0.25',
            'randomize_interval': '0',
            'export_whole_frame': '0',
            'export_crops': '1',
            'crop_size': '640',
            'random_offset_range': '600',
            'filename_prefix': ''
        }
        config['Workflow settings'] = {
            'Scan_default_folders': '1'
        }

        # Check if settings_crop.ini exists, and create it with default values if not
        if not os.path.exists('settings_crop.ini'):
            with open('settings_crop.ini', 'w', encoding='utf-8') as configfile:
                config.write(configfile)
        return config

    def config_read(self):
        logger = self.logger
        logger.debug('Running function config_read()')

        try:
            # Read settings from settings_crop.ini
            self.config.read('settings_crop.ini', encoding='utf-8')

            # Get values from the config file
            self.ocr_tesseract_path = self.config['Resource Paths'].get('OCR_tesseract_path',
                                                                        'C:/Program Files/Tesseract-OCR/tesseract.exe').strip()
            self.video_folder_path = self.config['Resource Paths'].get('video_folder_path', '').strip()
            self.annotation_file_path = self.config['Resource Paths'].get('annotation_file_path', '').strip()
            self.output_folder = self.config['Resource Paths'].get('output_folder', 'output').strip()
            self.scan_folders = self.config['Workflow settings'].get('Scan_default_folders', '0').strip()

            # Get crop values from config
            self.crop_mode = int(self.config['Crop settings'].get('crop_mode', '1').strip())
            self.frame_skip = int(self.config['Crop settings'].get('crop_interval_frames', '30').strip())
            self.frames_per_visit = int(self.config['Crop settings'].get('frames_per_visit', '0').strip())
            self.filter_visitors = int(self.config['Crop settings'].get('filter_visitors', '0').strip())
            self.yolo_processing = int(self.config['Crop settings'].get('yolo_processing', '0').strip())
            self.default_label_category = int(self.config['Crop settings'].get('default_label_category', '6').strip())
            self.yolo_conf = float(self.config['Crop settings'].get('yolo_conf', '0.25').strip())
            self.randomize = int(self.config['Crop settings'].get('randomize_interval', '0').strip())
            self.whole_frame = int(self.config['Crop settings'].get('export_whole_frame', '0').strip())
            self.cropped_frames = int(self.config['Crop settings'].get('export_crops', '1').strip())
            self.crop_size = int(self.config['Crop settings'].get('crop_size', '640').strip())
            self.offset_range = int(self.config['Crop settings'].get('random_offset_range', '600').strip())
            self.prefix = self.config['Crop settings'].get('filename_prefix', '').strip()

        except ValueError:
            logger.info('Error: Invalid folder/file path or crop settings found in settings_crop.ini')

    def load_videos(self):
        logger = self.logger
        logger.debug('Running function load_videos()')

        # Check if the video folder path is valid
        if utils.check_path(self.video_folder_path, 0):
            # Load videos
            try:
                self.video_filepaths = [
                    os.path.join(self.video_folder_path, f)
                    for f in os.listdir(self.video_folder_path)
                    if f.endswith('.mp4')
                ]
            except OSError as e:
                logger.error(f"Error: Failed to load videos: {e}")
                self.video_filepaths = []
        else:
            messagebox.showerror("Error", "Invalid video folder path")
            self.video_filepaths = []

    def reload_points_of_interest(self):
        logger = self.logger
        logger.debug('Running function reload_points_of_interest()')

        # Clear the array of POIs and reconstruct it with empty lists.
        self.points_of_interest_entry.clear()
        self.points_of_interest_entry = [[] for _ in range(len(self.video_filepaths))]

    def load_video_frames(self):
        # Define logger
        logger = self.logger
        logger.debug(f"Running function load_video_frames()")

        # Loop through each file in folder
        self.frames = []
        if utils.check_path(self.video_folder_path, 0):
            for filename in os.listdir(self.video_folder_path):
                if filename.endswith(".mp4"):  # Modify file extension as needed
                    video_path = os.path.join(self.video_folder_path, filename)
                    try:
                        # Use OpenCV or other library to extract first frame of video
                        # and add it to the frames list
                        with cv2.VideoCapture(video_path) as cap:
                            cap.set(cv2.CAP_PROP_POS_FRAMES, 25)
                            ret, frame = cap.read()
                            if ret:
                                self.frames.append(frame)
                            else:
                                # If the read operation fails, add a default image
                                default_image = cv2.imread('resources/img/nf.png')
                                self.frames.append(default_image)
                    except (cv2.error, OSError) as e:
                        logger.error(f"Error: Failed to process video '{filename}': {e}")
                        default_image = cv2.imread('resources/img/nf.png')
                        self.frames.append(default_image)
        else:
            logger.error("Error: Invalid video folder path")
            messagebox.showerror("Error", "Invalid video folder path")

######################################### SECTION DEALS WITH THE BACKEND ###############################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

######################################### EXTRACT TIME FROM VIDEO ######################################################
    def get_video_start_end_times(self, video_filepath):

        # Define logger
        self.logger.debug(f"Running function get_video_start_end_times({video_filepath})")
        video_filename = os.path.basename(video_filepath)
        self.logger.info(' '.join(["Flow:", "Processing video file -", video_filepath]))

        # get start time
        # get the time from filename
        parts = video_filename[:-4].split("_")
        if len(parts) == 6:
            start_time_minutes = "_".join([parts[3], parts[4], parts[5]])
            self.logger.info(' '.join(
                ["Flow: Video name format with prefixes detected. Extracted the time values -", start_time_minutes]))
        else:
            self.logger.info(
                "Error: Some video file names have an unsupported format. Expected format is "
                "CO_LO1_SPPSPP1_YYYYMMDD_HH_MM. Script assumes format YYYYMMDD_HH_MM.")
            start_time_minutes = video_filename[:-4]
        start_time_seconds, success = self.get_metadata_from_video(video_filepath, "start")
        if not success:
            text, frame = self.get_text_from_video(video_filepath, "start")
            start_time_seconds, success = self.process_OCR_text(text, frame, video_filepath, "start")
            if not success:
                start_time_seconds, success = self.get_text_manually(frame)
        start_time_str = '_'.join([start_time_minutes, start_time_seconds])
        start_time = pd.to_datetime(start_time_str, format='%Y%m%d_%H_%M_%S')

        # get end time
        end_time_seconds, success = self.get_metadata_from_video(video_filepath, "end")
        if not success:
            text, frame = self.get_text_from_video(video_filepath, "end")
            end_time_seconds, success = self.process_OCR_text(text, frame, video_filepath, "end")
            if not success:
                end_time_seconds, success = self.get_text_manually(frame)
        try:
            parser = createParser(video_filepath)
            metadata = extractMetadata(parser)
            duration = str(metadata.get("duration"))
            time_parts = duration.split(":")
            delta = int(time_parts[1])
        except:
            delta = 15 + (int(end_time_seconds) // 60)
        # print(start_time_minutes)
        # print(end_time_seconds)
        end_time_seconds = str(int(end_time_seconds) % 60)
        end_time_str = pd.to_datetime('_'.join([start_time_minutes, end_time_seconds]), format='%Y%m%d_%H_%M_%S')
        end_time = end_time_str + pd.Timedelta(minutes=int(delta))
        return start_time, end_time

    def get_metadata_from_video(self, video_filepath, start_or_end):
        if start_or_end == "start":
            try:
                parser = createParser(video_filepath)
                metadata = extractMetadata(parser)
                modify_date = str(metadata.get("creation_date"))
                return_time = modify_date[-2:]
                print("Flow: Obtained video start time from metadata.")
                success = True
            except:
                success = False
        elif start_or_end == "end":
            try:
                parser = createParser(video_filepath)
                metadata = extractMetadata(parser)
                modify_date = str(metadata.get("creation_date"))
                start_seconds = int(modify_date[-2:])
                duration = str(metadata.get("duration"))
                time_parts = duration.split(":")
                seconds = int(time_parts[2].split(".")[0])
                return_time = str(seconds + start_seconds)
                print("Flow: Obtained video end time from metadata.")
                success = True
            except:
                success = False
        return return_time, success

    def get_text_from_video(self, video_filepath, start_or_end):

        #Define logger
        self.logger.debug(f'Running function get_text_from_video({video_filepath}, {start_or_end})')

        # Define config
        config = self.config

        # Read settings from settings_crop.ini
        config.read('settings_crop.ini', encoding='utf-8')
        try:
            self.x_coordinate = int(config['OCR settings'].get('x_coordinate', '0').strip())
            self.y_coordinate = int(config['OCR settings'].get('y_coordinate', '0').strip())
            self.width = int(config['OCR settings'].get('width', '500').strip())
            self.height = int(config['OCR settings'].get('height', '40').strip())
        except ValueError:
            # Handle cases where conversion to integer fails
            self.logger.info('Error: Invalid integer value found in settings_crop.ini')

        # Get the video capture and ROI defined
        self.cap = cv2.VideoCapture(video_filepath)
        self.text_roi = (self.x_coordinate, self.y_coordinate, self.width, self.height)  # x, y, width, height

        # Define which frame to scan - start of end?
        if start_or_end == "end":
            total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            second_to_last_frame_idx = total_frames - 5
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, second_to_last_frame_idx)
        else:
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, 24)

        # Get the video frame and process the image
        ret, frame = self.cap.read()
        if ret:
            # Crop the image and pre-process it
            # height, width, channels = frame.shape
            x, y, w, h = self.text_roi
            text_frame = frame[y:y + h, x:x + w]
            HSV_img = cv2.cvtColor(text_frame, cv2.COLOR_BGR2HSV)
            h, s, v = cv2.split(HSV_img)
            v = cv2.GaussianBlur(v, (1, 1), 0)
            thresh = cv2.threshold(v, 220, 255, cv2.THRESH_BINARY_INV)[
                1]  # change the second number to change the threshold - anything over that value will be turned into white
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, ksize=(1, 1))
            thresh = cv2.dilate(thresh, kernel)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, ksize=(1, 1))
            thresh = cv2.erode(thresh, kernel)
            # text recognition
            OCR_text = pytesseract.image_to_string(thresh)

            # debug OCR file creation
            script_dir = os.path.dirname(os.path.abspath(__file__))
            output_dir = os.path.join(script_dir, "OCR images")
            utils.create_dir(output_dir)
            OCR_file_name = ''.join([os.path.basename(video_filepath)[:-4], "_", start_or_end, ".png"])
            OCR_file_path = os.path.join(output_dir, OCR_file_name)
            # cv2.imwrite(OCR_file_path, thresh)
            with open(OCR_file_path, 'wb') as f:
                f.write(cv2.imencode('.png', thresh)[1].tobytes())
        else:
            OCR_text = "none"
        self.cap.release()
        return OCR_text, frame

    def process_OCR_text(self, detected_text, frame, video_filepath, start_or_end):

        # Define logger
        self.logger.debug(f'Running function process_OCR_text({detected_text}, {video_filepath}, {start_or_end})')

        # Define variables
        return_time = "00"
        if "\n" in detected_text:
            detected_text = detected_text.replace("\n", "")
        if not len(detected_text) <= 23:
            detected_text = detected_text.rstrip()
            while not detected_text[-1].isdigit():
                detected_text = detected_text[:-1]
        correct_format = r"(0[0-9]|[1-5][0-9]):(0[0-9]|[1-5][0-9]):(0[0-9]|[1-5][0-9])"
        if re.match(correct_format, detected_text[-8:]):
            self.logger.info(' '.join(["Flow:", "Text detection successful -", detected_text[-8:]]))
            return_time = detected_text[-2:]
            success = True
        else:
            self.logger.info(' '.join(["Flow:", "Text detection failed -", detected_text]))
            success = False
        return return_time, success

    def get_text_manually(self, frame):

        def submit_time(input_field):
            self.logger.debug(f'Running function submit_time({input_field})')

            # Define variables
            text = input_field.get()
            dialog = self.manual_text_input_window

            # Validate text
            if not text.isdigit() or len(text) != 2 or int(text) > 59:
                # execute code here for when text is not in "SS" format
                self.logger.info(
                    "Error: OCR detected text does not follow the expected format. Manual input is not in the correct format. The value will be set to an arbitrary 00.")
                text = '00'
            else:
                self.logger.info("Error: OCR detected text does not follow the expected format. Resolved manually.")
            self.sec_OCR = text
            while True:
                try:
                    if dialog.winfo_exists():
                        dialog.quit()
                        dialog.destroy()
                        break
                except:
                    time.sleep(0.1)
                    break

        # Define variables
        root = self.main_window

        # Loop until the main window is created
        while True:
            try:
                if root.winfo_exists():
                    break
            except:
                time.sleep(0.1)

        # Create the dialog window
        dialog = tk.Toplevel(root)
        self.manual_text_input_window = dialog
        try:
            screen_width = dialog.winfo_screenwidth()
            screen_height = dialog.winfo_screenheight()
        except:
            screen_width = 1920
            screen_height = 1080

        dialog.wm_attributes("-topmost", 1)
        dialog.title("Time Input")

        # convert frame to tkinter image
        self.text_roi = (self.x_coordinate, self.y_coordinate, self.width, self.height)
        x, y, w, h = self.text_roi
        img_frame_width = min(screen_width // 2, w * 2)
        img_frame_height = min(screen_height // 2, h * 2)
        frame = frame[y:y + h, x:x + w]
        img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = cv2.resize(img, (img_frame_width, img_frame_height), Image.LANCZOS)
        img = Image.fromarray(img)
        img = ImageTk.PhotoImage(img)
        img_width = img.width()
        img_height = img.height()

        # Add image frame containing the video frame
        img_frame = tk.Frame(dialog, width=img_width, height=img_height)
        img_frame.pack(side=tk.TOP, pady=(0, 0))
        img_frame.pack_propagate(0)

        # Add image to image frame
        img_label = tk.Label(img_frame, image=img)
        img_label.pack(side=tk.TOP, pady=(0, 0))

        # Add label
        text_field = tk.Text(dialog, height=2, width=120, font=("Arial", 10))
        text_field.insert(tk.END,
                          "The OCR detection apparently failed.\nEnter the last two digits of the security camera watermark (number of seconds).\nThis will ensure cropping will happen at the right times")
        text_field.configure(state="disabled", highlightthickness=1, relief="flat", background=dialog.cget('bg'))
        text_field.tag_configure("center", justify="center")
        text_field.tag_add("center", "1.0", "end")
        text_field.pack(side=tk.TOP, padx=(0, 0))
        label = tk.Label(dialog, text="Enter text", font=("Arial", 10), background=dialog.cget('bg'))
        label.pack(pady=2)

        # Add input field
        j = 0
        input_field = tk.Entry(dialog, font=("Arial", 10), width=4)
        input_field.pack(pady=2)
        input_field.bind("<Return>", lambda j=j: submit_time(input_field))
        input_field.focus()

        # Add submit button
        submit_button = tk.Button(dialog, text="Submit", font=("Arial", 10),
                                  command=lambda j=j: submit_time(input_field))
        submit_button.pack(pady=2)

        # Position the window
        dialog_width = dialog.winfo_reqwidth() + img_frame_width
        dialog_height = dialog.winfo_reqheight() + img_frame_height
        dialog_pos_x = int((screen_width // 2) - (img_frame_width // 2))
        dialog_pos_y = 0
        dialog.geometry(f"{dialog_width}x{dialog_height}+{dialog_pos_x}+{dialog_pos_y}")

        # Start the dialog
        dialog.mainloop()

        # When dialog is closed
        return_time = self.sec_OCR
        if len(return_time) > 0:
            success = True
        else:
            success = False
        return return_time, success

######################################### EXTRACT DATA FROM EXCEL ######################################################

    def evaluate_string_formula(self, cell):

        # If the cell contains a number, return the value as is
        if isinstance(cell, (int, float)):
            return cell
        # If the cell contains an Excel formula, use openpyxl to evaluate it
        elif cell.startswith('='):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'].value = cell
            value = ws['A1'].value
            wb.close()
            return value
        # If the cell contains text, return the text as is
        else:
            return cell

    def load_excel_table(self, file_path):

        # The default values of cols to be extracted are:
        # 0 - A - Year
        # 1 - B - Month - !
        # ...
        # 5 - F - Seconds
        # 15 - P - Visitor arrival - filter column
        # 18 - S - Visit duration in seconds
        # 19 - T - Time of departure - Hours
        # 20 - U - Time of departure - Minutes
        # 21 - V - Time of departure - Seconds
        # 23 - X - Insect species
        # 24 - Y - Insect Order

        # Define logger
        self.logger.debug(f"Running function load_excel_table({file_path})")

        # Read the Excel file, skipping the first two rows
        cols: list[int] = [0, 1, 2, 3, 4, 5, 15, 18, 19, 20, 21, 23, 24]
        converters: Dict[int, Callable] = {0: self.evaluate_string_formula, 1: self.evaluate_string_formula,
                                           2: self.evaluate_string_formula, 3: self.evaluate_string_formula,
                                           4: self.evaluate_string_formula, 18: self.evaluate_string_formula}
        try:
            df = pd.read_excel(file_path, usecols=cols, skiprows=2, header=None,
                               converters=converters)
        except ValueError as e:
            self.logger.error(f"Error reading Excel file {file_path}. Error message: {e}")

            # Open the Excel workbook using xlwings
            wb = xw.Book(file_path)
            sheet = wb.sheets[0]

            # Remove any filters
            if sheet.api.AutoFilterMode:
                sheet.api.AutoFilterMode = False

            # Save to temporary file
            utils.create_dir("resources/exc")
            temp_file_path = os.path.join("resources/exc", "temp.xlsx")
            wb.save(temp_file_path)
            wb.close()

            # Read with pandas
            try:
                df = pd.read_excel(temp_file_path, usecols=cols, skiprows=2, header=None,
                                   converters=converters)
            except ValueError as e:
                self.logger.error(
                    f"Attempted to fix errors in Excel file {file_path}. Attempt failed. Error message: {e}. Please fix the errors manually and try again.")
                return None
            self.logger.info(
                f"Attempted to remove filters from Excel file {file_path}. Saved a copy of the file to {temp_file_path}")

        self.logger.info(f"Flow: Retrieved dataframe from Excel:\n{df}")

        # Filter data frame based on whether the value in the column of index 6 (P - visitor arrival) is 1.
        col_to_filter_by: int = 6
        filtered_df = df[df.iloc[:, col_to_filter_by] == 1]

        # Convert year to integer
        col_year = 0
        filtered_df.iloc[:, col_year] = filtered_df.iloc[:, col_year].astype(int)

        # Convert the month abbreviations in column 2 to month numbers
        col_month = 1
        months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                  'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        filtered_df.iloc[:, col_month] = filtered_df.iloc[:, col_month].replace(months)
        filtered_df.iloc[:, col_month] = filtered_df.iloc[:, col_month].astype(int)
        filtered_df = filtered_df.copy()

        # Convert integers of hours/minutes/seconds into 02 format.
        cols_time = [2,3,4,5,6]
        for i in range(len(cols_time)):
            j = cols_time[i]
            filtered_df.iloc[:, j] = filtered_df.iloc[:, j].astype(int).apply(lambda x: f'{x:02}')

        # Add another column and move the 11th
        num_columns = filtered_df.shape[1]
        filtered_df.loc[:, num_columns] = filtered_df.iloc[:, 11]
        filtered_df.iloc[:, 12] = filtered_df.iloc[:, 12]
        filtered_df.iloc[:, 11] = filtered_df.iloc[:, 0:6].apply(lambda x: f"{x[0]}{x[1]}{x[2]}_{x[3]}_{x[4]}_{x[5]}", axis=1)

        # Debug check
        #print(f"Flow: Filtered dataframe:\n {filtered_df}")

        # Take visit duration and the assembled timestamp
        filtered_data = filtered_df.iloc[:, [7, 11]].values.tolist()


        # Get the column names
        column_1 = filtered_df.columns[12]
        column_2 = filtered_df.columns[13]

        # Count the number of NAs in each column
        na_count_1 = filtered_df[column_1].isna().sum()
        na_count_2 = filtered_df[column_2].isna().sum()

        # Check which column has fewer NAs
        if na_count_1 <= na_count_2:
            chosen_column = column_1
            other_column = column_2
        else:
            chosen_column = column_2
            other_column = column_1

        # Replace NAs in chosen_column with values from other_column if they are not NAs
        visitor_id = filtered_df[[chosen_column, other_column]].copy()
        visitor_id[chosen_column].fillna(visitor_id[other_column], inplace=True)
        visitor_id = visitor_id[[chosen_column]].values.tolist()
        #print(visitor_id)

        annotation_data_array = filtered_data
        utils.create_dir("resources/exc/")
        filtered_df.to_excel("resources/exc/output_filtered_crop.xlsx", index=False)
        return annotation_data_array, visitor_id

    def load_csv(self, file_path):

        # Define logger
        self.logger.debug(f"Running function load_csv({file_path})")

        # Define the columns to extract
        cols: list[int] = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        converters: Dict[int, Callable] = {0: self.evaluate_string_formula, 1: self.evaluate_string_formula,
                                                    2: self.evaluate_string_formula,
                                                    3: self.evaluate_string_formula, 4: self.evaluate_string_formula,
                                                    6: self.evaluate_string_formula}

        # Read the Excel file, skipping the first two rows - follow the custom format
        try:
            filtered_df = pd.read_excel(file_path, usecols=cols, skiprows=2, header=None,
                                        converters=converters)
        except ValueError as e:
            self.logger.error(f"Error reading Excel file {file_path}. Error message: {e}")
            # Open the Excel workbook using xlwings
            wb = xw.Book(file_path)

            sheet = wb.sheets[0]

            # Remove any filters
            if sheet.api.AutoFilterMode:
                sheet.api.AutoFilterMode = False

            # Save to temporary file
            utils.create_dir("resources/exc")
            temp_file_path = os.path.join("resources/exc", "temp.xlsx")
            wb.save(temp_file_path)
            wb.close()

            # Read with pandas
            try:
                filtered_df = pd.read_excel(file_path, usecols=cols, skiprows=2, header=None,
                                            converters=converters)
            except ValueError as e:
                self.logger.error(
                    f"Attempted to fix errors in Excel file {file_path}. Attempt failed. Error message: {e}. Please fix the errors manually and try again.")
                return None
            self.logger.info(
                f"Attempted to remove filters from Excel file {file_path}. Saved a copy of the file to {temp_file_path}")
        self.logger.info(f"Flow: Retrieved dataframe from Excel:\n{filtered_df}")

        # Convert the month abbreviations in column 2(1) to month numbers
        months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                  'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        filtered_df.iloc[:, 1] = filtered_df.iloc[:, 1].replace(months)
        filtered_df.iloc[:, 1] = filtered_df.iloc[:, 1].astype(int)
        filtered_df = filtered_df.copy()
        for i in range(5):
            j = i + 1
            filtered_df.iloc[:, j] = filtered_df.iloc[:, j].astype(int).apply(lambda x: f'{x:02}')
        filtered_df.loc[:, 10] = filtered_df.iloc[:, 0:6].apply(lambda x: f"{x[0]}{x[1]}{x[2]}_{x[3]}_{x[4]}_{x[5]}",
                                                                axis=1)
        filtered_df.iloc[:, 0] = filtered_df.iloc[:, 0].astype(int)
        self.logger.info(f"Flow: Filtered dataframe:\n {filtered_df}")

        # convert to list
        filtered_data = filtered_df.iloc[:, [6, 10]].values.tolist()

        annotation_data_array = filtered_data
        utils.create_dir("resources/exc/")
        filtered_df.to_excel("resources/exc/output_filtered_crop.xlsx", index=False)
        return annotation_data_array

    ######################################### CROP FUNCTIONALITY ######################################################

    def get_video_data(self, video_filepaths):

        # Define logger
        self.logger.debug(f"Running function get_video_data({video_filepaths})")

        # loop through time annotations and open corresponding video file
        # extract video data beforehand to save processing time
        video_data = []
        i: int
        for i, filepath in enumerate(video_filepaths):
            if filepath.endswith('.mp4'):
                video_start_time, video_end_time = self.get_video_start_end_times(video_filepaths[i])
                video_data_entry = [video_filepaths[i], video_start_time, video_end_time]
                video_data.append(video_data_entry)
        return video_data

    async def capture_crop(self, frame, point):

       # Define logger
        self.logger.debug(f"Running function capture_crop({point})")

        # Prepare local variables
        x, y = point

        # Add a random offset to the coordinates, but ensure they remain within the image bounds
        frame_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        frame_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

        # Check if any of the dimensions is smaller than crop_size and if so upscale the image to prevent crops smaller than desired crop_size
        if frame_height < self.crop_size or frame_width < self.crop_size:
            # Calculate the scaling factor to upscale the image
            scaling_factor = self.crop_size / min(frame_height, frame_width)

            # Calculate the new dimensions for the upscaled frame
            new_width = int(round(frame_width * scaling_factor))
            new_height = int(round(frame_height * scaling_factor))

            # Upscale the frame using cv2.resize with Lanczos upscaling algorithm
            frame = cv2.resize(frame, (new_width, new_height), interpolation=cv2.INTER_LANCZOS4)

        # Get the new frame size
        frame_height, frame_width = frame.shape[:2]

        # Calculate the coordinates for the area that will be cropped
        x_offset = random.randint(-self.offset_range, self.offset_range)
        y_offset = random.randint(-self.offset_range, self.offset_range)
        x1 = max(0, min(((x - self.crop_size // 2) + x_offset), frame_width - self.crop_size))
        y1 = max(0, min(((y - self.crop_size // 2) + y_offset), frame_height - self.crop_size))
        x2 = max(self.crop_size, min(((x + self.crop_size // 2) + x_offset), frame_width))
        y2 = max(self.crop_size, min(((y + self.crop_size // 2) + y_offset), frame_height))

        # Crop the image
        crop = frame[y1:y2, x1:x2]

        # Convert to correct color space
        crop_img = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
        if crop_img.shape[2] == 3:
            crop_img = cv2.cvtColor(crop_img, cv2.COLOR_BGR2RGB)

        # Return the cropped image and the coordinates for future reference
        return crop_img, x1, y1, x2, y2

    async def generate_frames(self, frame, success, tag, index, frame_number_start):

        # Define logger
        self.logger.debug(f"Running function generate_frames({index})")

        # Prepare name elements
        species = tag[-27:-19].replace("_", "")
        timestamp = tag[-18:-4]

        # Define local variables
        crop_counter = 1
        frame_skip_loc = self.frame_skip

        # Calculate the frame skip variable based on the limited number of frames per visit
        if self.frames_per_visit > 0:
            frame_skip_loc = int((self.visit_duration * self.fps) // self.frames_per_visit)
            if frame_skip_loc < 1:
                frame_skip_loc = 1

        # Loop through the video and crop y images every n-th frame
        frame_count = 0
        img_paths = []

        while success:
            # Crop images every n-th frame
            if int(frame_count % frame_skip_loc) == 0:
                for i, point in enumerate(self.points_of_interest_entry[index]):
                    if self.cropped_frames == 1:
                        crop_img, x1, y1, x2, y2 = await self.capture_crop(frame, point)
                        img_path = f"./{self.output_folder}/{self.prefix}{species}_{timestamp}_{frame_number_start + frame_count}_{crop_counter}_{i + 1}_{x1},{y1}_{x2},{y2}.jpg"
                        cv2.imwrite(img_path, crop_img)
                        img_paths.append(img_path)
                if self.whole_frame == 1:
                    img_path = f"./{self.output_folder}/whole frames/{self.prefix}{species}_{timestamp}_{frame_number_start + frame_count}_{crop_counter}_whole.jpg"
                    cv2.imwrite(img_path, frame)
                crop_counter += 1

            # If the random frame skip interval is activated add a random number to the counter or add the set frame skip interval
            if self.randomize == 1:
                if (frame_skip_loc - frame_count == 1):
                    frame_count += 1
                else:
                    frame_count += random.randint(1, max((frame_skip_loc - frame_count), 2))
            else:
                frame_count += frame_skip_loc

            # Read the next frame
            frame_to_read = frame_number_start + frame_count
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_to_read)
            success, frame = self.cap.read()

            # If the frame count is equal or larger than the amount of frames that comprises the duration of the visit end the loop
            if not (frame_count <= (self.visit_duration * self.fps)):
                # Release the video capture object and close all windows
                self.cap.release()
                cv2.destroyAllWindows()
                break

        # Return the resulting list of image paths for future reference
        return img_paths

    async def yolo_preprocessing(self, img_paths, valid_annotations_array, index):
        model = YOLO('resources/yolo/best.pt')
        results = model(img_paths, save=False, imgsz=self.crop_size, conf=self.yolo_conf, save_txt=False, max_det=1, stream=True)
        for i, result in enumerate(results):
            boxes = result.boxes.data
            original_path = os.path.join(img_paths[i])
            utils.create_dir(f"{self.output_folder}/empty")
            utils.create_dir(f"{self.output_folder}/visitor")
            utils.create_dir(f"{self.output_folder}/visitor/labels")
            empty_path = os.path.join(f"{self.output_folder}/empty", os.path.basename(img_paths[i]))
            visitor_path = os.path.join(f"{self.output_folder}/visitor", os.path.basename(img_paths[i]))
            label_path = os.path.join(f"{self.output_folder}/visitor/labels", os.path.basename(img_paths[i])[:-4])
            if len(result.boxes.data) > 0:
                shutil.move(original_path, visitor_path)
                with open(f"{label_path}.txt", 'w') as file:
                    # Write the box_data to the fil
                    txt = []
                    lst = result.boxes.xywhn[0].tolist()
                    for item in lst:
                        txt_item = round(item, 6)
                        txt.append(txt_item)
                    txt = str(txt)
                    if any(len(row) < 7 for row in valid_annotations_array):
                        visitor_category = self.default_label_category
                    else:
                        visitor_category = valid_annotations_array[index][6]
                    file.write(f"{visitor_category} {txt.replace('[', '').replace(']', '').replace(',', '')}")
            else:
                shutil.move(original_path, empty_path)

    def crop_engine(self):

        # Define logger
        self.logger.debug("Running function crop_engine()")

        # Define variables
        root = self.main_window

        # Ask to confirm whether the process should begin
        result = utils.ask_yes_no("Do you want to start the cropping process?")
        if result:
            if len(self.points_of_interest_entry[0]) == 0:
                messagebox.showinfo("Warning", "No points of interest selected. Please select at least one POI.")
                self.reload(1, False)
                return
            valid_annotations_array = []
            valid_annotation_data_entry = []
            print(f"Flow: Start cropping on the following videos: {self.video_filepaths}")
            root.withdraw()
            if not self.crop_mode == 3:
                video_ok = utils.check_path(self.video_folder_path, 0)
                excel_ok = utils.check_path(self.annotation_file_path, 1)
                if not video_ok or not excel_ok:
                    messagebox.showinfo("Warning",
                                        f"Unspecified path to a video folder or a valid Excel file.")
                    self.reload(1, False)
                    return
                if self.crop_mode == 1:
                    video_data = get_video_data(self.video_filepaths)
                    annotation_data_array, visitor_id = load_excel_table(self.annotation_file_path)
                if self.crop_mode == 2:
                    annotation_data_array = load_csv(self.annotation_file_path)
                if annotation_data_array is None:
                    messagebox.showinfo("Warning",
                                        f"Attempted to fix errors in the selected Excel file. Attempt failed. Please fix the errors manually and try again.")
                    self.reload(1, False)
                    return
                if self.crop_mode == 2:
                    video_filepaths_temp = self.video_filepaths
                    self.video_filepaths = []
                    for index, list in enumerate(annotation_data_array):
                        for filepath in video_filepaths_temp:
                            filename = os.path.basename(filepath)  # get just the filename from the full path
                            if annotation_data_array[index][1][:-9] in filename:
                                if datetime.timedelta() <= datetime.datetime.strptime(
                                        annotation_data_array[index][1][-8:-3], '%H_%M') - datetime.datetime.strptime(
                                    filename[-9:-4], '%H_%M') <= datetime.timedelta(minutes=15):
                                    self.video_filepaths.append(filepath)
                    video_data = self.get_video_data(self.video_filepaths)
                print(f"Flow: Start cropping according to the following annotations: {annotation_data_array}")
                print(f"Flow: Start cropping with the following video data: {video_data}")
                for index, list in enumerate(annotation_data_array):
                    print(' '.join(["Flow: Annotation number:", str(index + 1)]))
                    annotation_time = pd.to_datetime(annotation_data_array[index][1], format='%Y%m%d_%H_%M_%S')
                    for i, list in enumerate(video_data):
                        print(f"{video_data[i][1]} <= {annotation_time} <= {video_data[i][2]}")
                        if video_data[i][1] <= annotation_time <= video_data[i][2]:
                            for each in range(len(annotation_data_array[index])):
                                valid_annotation_data_entry.append(annotation_data_array[index][each])
                            for each in range(3):
                                valid_annotation_data_entry.append(video_data[i][each])
                            if self.crop_mode == 1:
                                valid_annotation_data_entry.append(visitor_id[index][0])
                            print(f"Flow: Relevant annotations: {valid_annotation_data_entry}")
                            valid_annotations_array.append(valid_annotation_data_entry)
                            valid_annotation_data_entry = []
                if self.filter_visitors == 1:
                    filter_array_by_visitors(valid_annotations_array)
                    if len(self.filtered_array) > 0:
                        valid_annotations_array = self.filtered_array
                    else:
                        print("No visitors of the selected type found.")
                        self.reload(1, False)
                        return
                for index in range(len(valid_annotations_array)):
                    print(f"Processing item {index}")
                    self.visit_index = index
                    annotation_time = pd.to_datetime(valid_annotations_array[index][1], format='%Y%m%d_%H_%M_%S')
                    annotation_offset = (annotation_time - valid_annotations_array[index][3]).total_seconds()
                    self.cap = cv2.VideoCapture(valid_annotations_array[index][2])
                    total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
                    self.fps = self.cap.get(cv2.CAP_PROP_FPS)
                    self.visit_duration = (min(((annotation_offset * self.fps) + (int(valid_annotations_array[index][0]) * self.fps)),
                                          total_frames) - (annotation_offset * self.fps)) // self.fps
                    frame_number_start = int(annotation_offset * self.fps)
                    self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number_start)
                    success, frame = self.cap.read()
                    img_paths = asyncio.run(
                        self.generate_frames(frame, success, os.path.basename(valid_annotations_array[index][2]),
                                        self.video_filepaths.index(valid_annotations_array[index][2])))
                    # loop.close()
                    # img_paths = generate_frames(frame, success, os.path.basename(valid_annotations_array[index][2]),
                    # video_filepaths.index(valid_annotations_array[index][2]))
                    if self.yolo_processing == 1:
                        asyncio.run(self.yolo_preprocessing(img_paths, valid_annotations_array, index))
            else:
                video_ok = utils.check_path(self.video_folder_path, 0)
                if not video_ok:
                    messagebox.showinfo("Warning",
                                        f"Unspecified path to a video folder.")
                    self.reload(1, False)
                    return
                orig_wf = self.whole_frame
                self.whole_frame = 1
                for i, filepath in enumerate(self.video_filepaths):
                    self.cap = cv2.VideoCapture(self.video_filepaths[i])
                    self.fps = self.cap.get(cv2.CAP_PROP_FPS)
                    self.visit_duration = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT)) // self.fps
                    frame_number_start = 2
                    self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number_start)
                    success, frame = self.cap.read()
                    img_paths = self.generate_frames(frame, success, os.path.basename(self.video_filepaths[i]), i)
                self.whole_frame = orig_wf
        self.reload(1, False)
        # try:
        #     if root.winfo_exists():
        #         root.destroy()
        # except:
        #     print("Error: Unexpected, window destroyed before reference. Odpruženo.")
        # loaded = 1
        # open_ICCS_window()

######################################### SECTION DEALS WITH THE GUI ###################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

    def load_icon(self, path, size: tuple = (50, 50)):

        stream = open(path, "rb")
        bytes = bytearray(stream.read())
        numpyarray = np.asarray(bytes, dtype=numpy.uint8)
        img1 = cv2.imdecode(numpyarray, cv2.IMREAD_UNCHANGED)

        # Convert the BGR image to RGB (Tkinter uses RGB format)
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)

        # Convert the NumPy array to a PIL image
        pil_img = Image.fromarray(img1_rgb)

        # Create a Tkinter.PhotoImage from the PIL image
        img = ImageTk.PhotoImage(pil_img)

        #img = Image.open(path)
        #img = img.resize(size)
        #img = ImageTk.PhotoImage(img)

        self.gui_imgs.append(img)
        return img

    def open_main_window(self):

        # Define logger
        logger = self.logger
        logger.debug(f"Running function open_ICCS_window()")

        # Create tkinter window
        root = tk.Tk()
        self.main_window = root
        root.focus()
        root.title(f"{self.app_title} - Folder: {os.path.basename(os.path.normpath(self.video_folder_path))} - Table: {os.path.basename(os.path.normpath(self.annotation_file_path))}")

        # Arbitrary variable for lambdas
        j = 0

        # Create frame for the rest
        outer_frame = tk.Frame(root)
        outer_frame.pack(side=tk.TOP, expand=True, fill=tk.BOTH)

        top_toolbar = self.build_top_toolbar(outer_frame)
        center_frame = self.build_center_frame(outer_frame)
        top_toolbar.lift(center_frame)
        bottom_left_panel = self.build_bottom_left_panel(root)
        bottom_right_panel = self.build_bottom_right_panel(root)
        bottom_toolbar = self.build_bottom_toolbar(root)

        # Update button images
        if self.loaded == False:
            for each in range(len(self.video_filepaths)):
                first = -1
                update_button_image(self.frames[each].copy(), (max(each, 0) // 6), (each - ((max(each, 0) // 6) * 6)), first)
            # if auto is defined, set the auto processing to that value otherwise set it to 0
            if self.auto is None:
                self.auto_processing.set(0)
            else:
                self.auto_processing.set(self.auto)

        # Get the screen width and height
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        # Set the window size to fit the entire screen
        root.geometry(f"{screen_width}x{screen_height - 70}+-10+0")
        root.mainloop()

    def build_top_toolbar(self, parent):

        def can_switch_folder(where):
            btn_state = "disabled"
            if where == "right":
                if self.dir_hierarchy and (self.scanned_folders.index(os.path.basename(os.path.normpath(self.video_folder_path))) + 1) != len(self.scanned_folders):
                    btn_state = "normal"
                else:
                    btn_state = "disabled"
            elif where == "left":
                if self.dir_hierarchy and self.scanned_folders.index(os.path.basename(os.path.normpath(self.video_folder_path))) != 0:
                    btn_state = "normal"
                else:
                    btn_state = "disabled"
            return btn_state

        # Define variables
        outer_frame = parent
        j = 0

        # Create top toolbar
        toolbar = tk.Frame(outer_frame)
        toolbar.pack(side=tk.TOP, fill=tk.BOTH)

        # LEFT button
        left_button = tk.Button(toolbar, image=self.load_icon("resources/img/la.png"), compound=tk.LEFT, text="Previous folder", padx=10, pady=5,
                                height=48, width=200, state=can_switch_folder("left"), command=lambda j=j: switch_folder("left"))
        left_button.grid(row=0, column=0, padx=0, pady=5, sticky="ew")

        # MENU button
        menu_button = tk.Button(toolbar, image=self.load_icon("resources/img/mn.png"), compound=tk.LEFT, text="Menu", padx=10, pady=5, height=48,
                                command=lambda j=j: open_menu())
        menu_button.grid(row=0, column=1, padx=0, pady=5, sticky="ew")

        # frame for radio buttons
        radio_frame = tk.Frame(toolbar)
        radio_frame.grid(row=0, column=2, padx=0, pady=5, sticky="ew")

        # create a tkinter variable to hold the selected value
        selected_option = tk.StringVar(value=self.crop_mode)

        # CROP_MODE buttons
        # create the radio buttons and group them together
        rb1 = tk.Radiobutton(radio_frame, text="", image=self.load_icon("resources/img/1.png"), variable=selected_option, value=1, indicatoron=False,
                             height=56, width=116, font=("Arial", 17), command=lambda j_=j: update_crop_mode(1))
        rb2 = tk.Radiobutton(radio_frame, text="", image=self.load_icon("resources/img/2.png"), variable=selected_option, value=2, indicatoron=False,
                             height=56, width=116, font=("Arial", 17), command=lambda j=j: update_crop_mode(2))
        rb3 = tk.Radiobutton(radio_frame, text="", image=self.load_icon("resources/img/3.png"), variable=selected_option, value=3,
                             indicatoron=False,
                             height=56, width=116, font=("Arial", 17), command=lambda j=j: update_crop_mode(3))

        # arrange the radio buttons in a horizontal layout using the grid geometry manager
        rb1.grid(row=0, column=0, sticky="ew")
        rb2.grid(row=0, column=1, sticky="ew")
        rb3.grid(row=0, column=2, sticky="ew")
        radio_frame.grid_columnconfigure(0, weight=1, minsize=50)
        radio_frame.grid_columnconfigure(1, weight=1, minsize=50)
        radio_frame.grid_columnconfigure(2, weight=1, minsize=50)

        # AUTO checkbox
        on_image = tk.PhotoImage(width=116, height=57)
        off_image = tk.PhotoImage(width=116, height=57)
        on_image.put(("green",), to=(0, 0, 56, 56))
        off_image.put(("red",), to=(57, 0, 115, 56))
        self.auto_processing = tk.IntVar(value=0)
        self.auto_processing.set(0)
        cb1 = tk.Checkbutton(toolbar, image=off_image, selectimage=on_image, indicatoron=False, onvalue=1, offvalue=0,
                             variable=self.auto_processing)
        cb1.grid(row=0, column=3, padx=0, pady=5, sticky="ew")

        # AUTO button
        auto_button = tk.Button(toolbar, image=self.load_icon("resources/img/au.png"), compound=tk.LEFT, text="Automatic evaluation", padx=10,
                                pady=5, height=48,
                                command=lambda j=j: self.auto_processing.set(1 - self.auto_processing.get()))
        auto_button.grid(row=0, column=4, padx=0, pady=5, sticky="ew")

        # VIDEO FOLDER button
        fl_button = tk.Button(toolbar, image=self.load_icon("resources/img/fl.png"), compound=tk.LEFT, text="Select video folder", padx=10, pady=5,
                              height=48, command=lambda j=j: change_video_folder())
        fl_button.grid(row=0, column=5, padx=0, pady=5, sticky="ew")

        # EXCEL PATH button
        et_button = tk.Button(toolbar, image=self.load_icon("resources/img/et.png"), compound=tk.LEFT, text="Select Excel table", padx=10, pady=5,
                              height=48, command=lambda j=j: change_excel_path())
        et_button.grid(row=0, column=6, padx=0, pady=5, sticky="ew")

        # OCR button
        ocr_button = tk.Button(toolbar, image=self.load_icon("resources/img/ocr.png"), compound=tk.LEFT, text="OCR", padx=10, pady=5, height=48,
                               width=100, command=lambda j=j: set_ocr_roi(self.video_filepaths[0]))
        ocr_button.grid(row=0, column=7, padx=0, pady=5, sticky="ew")

        # RIGHT button
        right_button = tk.Button(toolbar, image=self.load_icon("resources/img/ra.png"), compound=tk.RIGHT, text="Next folder", padx=10, pady=5,
                                 height=48, width=200, state=can_switch_folder("right"), command=lambda j=j: switch_folder("right"))
        right_button.grid(row=0, column=8, padx=0, pady=5, sticky="ew")

        # configure columns of toolbox
        toolbar.grid_columnconfigure(0, weight=2, minsize=50)
        toolbar.grid_columnconfigure(1, weight=3, minsize=50)
        toolbar.grid_columnconfigure(2, weight=4, minsize=150)
        toolbar.grid_columnconfigure(3, weight=4, minsize=50)
        toolbar.grid_columnconfigure(4, weight=1, minsize=50)
        toolbar.grid_columnconfigure(5, weight=4, minsize=50)
        toolbar.grid_columnconfigure(6, weight=4, minsize=50)
        toolbar.grid_columnconfigure(7, weight=4, minsize=50)
        toolbar.grid_columnconfigure(8, weight=2, minsize=50)

        return toolbar

    def build_center_frame(self, parent):

        # Define variables
        outer_frame = parent

        # Create a canvas to hold the tile buttons
        canvas = tk.Canvas(outer_frame)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create a scrollbar for the canvas
        scrollbar = tk.Scrollbar(outer_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Configure the canvas to use the scrollbar
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Create target frame for the rest
        target_frame = tk.Frame(outer_frame)
        target_frame.pack(side=tk.TOP)

        self.button_images = []
        self.buttons = []
        rows = math.ceil(len(self.video_filepaths) / 6)
        # per_row = math.ceil(len(self.video_filepaths) // rows)
        label_frame = tk.Frame(target_frame, width=50)
        label_frame.pack(side=tk.LEFT)
        for i in range(rows):
            hour_1st = ((i) * 6)
            if (((i) * 6) + 6) + 1 <= len(self.video_filepaths):
                hour_2nd = max(((i) * 6) + 6, len(self.video_filepaths) - 1)
            else:
                hour_2nd = (len(self.video_filepaths) % 6) - 1
            text_label = tk.Label(label_frame,
                                  text=f"{os.path.basename(self.video_filepaths[hour_1st])[-9:-7]}-{os.path.basename(self.video_filepaths[hour_2nd])[-9:-7]}",
                                  font=("Arial", 15), background=outer_frame.cget('bg'))
            text_label.pack(side=tk.TOP, padx=30, pady=67)
        for i in range(rows):
            # Loop through first 24 frames and create buttons with images
            button_frame = tk.Frame(target_frame)
            button_frame.pack(side=tk.TOP)
            row = []
            for j in range(6):
                if (j + ((i) * 6)) < len(self.video_filepaths):
                    # Convert frame to PIL image
                    pil_img = cv2.cvtColor(self.frames[j + ((i) * 6)], cv2.COLOR_BGR2RGB)
                    pil_img = Image.fromarray(pil_img, mode='RGB')
                    # Resize image to fit button
                    pil_img = pil_img.resize((276, 156))
                    # Convert PIL image to tkinter image
                    tk_img = ImageTk.PhotoImage(pil_img)
                    self.button_images.append(tk_img)
                    # Create button with image and add to button frame
                    button = tk.Button(button_frame, image=self.button_images[j + ((i) * 6)],
                                       command=lambda i=i, j=j: on_button_click(i, j, self.button_images))
                    button.grid(row=i, column=j, sticky="w")
                    row.append(button)
                else:
                    # Create a dummy frame to fill in the grid
                    new_img = Image.new("RGBA", (276, 156), (0, 0, 0, 0))
                    new_img = ImageTk.PhotoImage(new_img)
                    dummy_frame = tk.Button(button_frame, image=new_img, foreground="white", state="disabled")
                    dummy_frame.grid(row=i, column=j, sticky='w')
                    row.append(dummy_frame)
            self.buttons.append(row)

        # Update the canvas to show the buttons
        canvas.create_window((0, 0), window=target_frame, anchor=tk.NW)
        target_frame.update_idletasks()

        # Configure the canvas to show the entire frame
        canvas.configure(scrollregion=canvas.bbox("all"))

        # Bind mouse wheel event to canvas
        canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(-1 * (event.delta // 120), "units"))

        return target_frame

    def build_bottom_left_panel(self, parent):

        # Define variables
        root = parent
        j = 0

        # Create panel frame
        bottom_left_panel = tk.Frame(root, pady=0)
        bottom_left_panel.pack(side=tk.LEFT)

        # Create buttons
        fl_button = tk.Button(bottom_left_panel, image=self.load_icon("resources/img/fl.png"), compound=tk.LEFT, text="", padx=10, pady=5, width=60,
                              height=58,
                              command=lambda j=j: os.startfile(self.video_folder_path))
        fl_button.pack(side=tk.LEFT)

        return bottom_left_panel

    def build_bottom_right_panel(self, parent):

        # Define variables
        root = parent
        j = 0

        # Create panel frame
        bottom_right_panel = tk.Frame(root, pady=0)
        bottom_right_panel.pack(side=tk.RIGHT)

        # Create buttons
        et_button = tk.Button(bottom_right_panel, image=self.load_icon("resources/img/et.png"), compound=tk.LEFT, text="", padx=10, pady=5, width=60,
                              height=58,
                              command=lambda j=j: os.startfile(self.annotation_file_path))
        et_button.pack(side=tk.RIGHT)

        ef_button = tk.Button(bottom_right_panel, image=self.load_icon("resources/img/ef.png"), compound=tk.LEFT, text="", padx=0, pady=5, width=60,
                              height=58,
                              command=lambda j=j: os.startfile(os.path.dirname(self.annotation_file_path)))
        ef_button.pack(side=tk.RIGHT)

        return bottom_right_panel

    def build_bottom_toolbar(self, parent):

        # Define variables
        root = parent
        j = 0

        # Create panel frame
        bottom_toolbar = tk.Frame(root, pady=5)
        bottom_toolbar.pack(side=tk.TOP, fill=tk.BOTH)

        # Create buttons
        save_button = tk.Button(bottom_toolbar, text="Save", image=self.load_icon("resources/img/sv_1.png"), compound=tk.LEFT, padx=10, pady=5,
                                width=300,
                                height=48, command=lambda j=j: save_progress())

        crop_button = tk.Button(bottom_toolbar, text="Crop", image=self.load_icon("resources/img/cr_1.png"), compound=tk.LEFT, padx=10, pady=5,
                                width=300,
                                height=48, command=lambda j=j: crop_engine())

        sort_button = tk.Button(bottom_toolbar, text="Sort", image=self.load_icon("resources/img/so.png"), compound=tk.LEFT, padx=10, pady=5,
                                width=300,
                                height=48, command=lambda j=j: sort_engine())

        load_button = tk.Button(bottom_toolbar, text="Load", image=self.load_icon("resources/img/lo.png"), compound=tk.LEFT, padx=10, pady=5,
                                width=300,
                                height=48, command=lambda j=j: load_progress())

        # Specify the column for each button
        save_button.grid(row=0, column=1, sticky="ew")
        crop_button.grid(row=0, column=2, sticky="ew")
        sort_button.grid(row=0, column=3, sticky="ew")
        load_button.grid(row=0, column=4, sticky="ew")

        # Add padding between the buttons
        bottom_toolbar.grid_columnconfigure(0, weight=1)
        bottom_toolbar.grid_columnconfigure(1, weight=2)
        bottom_toolbar.grid_columnconfigure(2, weight=2)
        bottom_toolbar.grid_columnconfigure(3, weight=2)
        bottom_toolbar.grid_columnconfigure(4, weight=2)
        bottom_toolbar.grid_columnconfigure(5, weight=1)

        return bottom_toolbar

    def reload(self, is_window_already_loaded: bool, reload_POIs: bool):

        # Define logger
        logger = self.logger
        logger.debug(f"Running function reload({is_window_already_loaded}, {reload_POIs})")

        # Define variables
        root = self.main_window

        # Reload videos, clear POIs, reload video frames
        self.load_videos()
        if reload_POIs:
            self.reload_points_of_interest()
        self.load_video_frames()
        self.close_main_window()
        self.loaded = is_window_already_loaded
        self.open_main_window()

######################################### SECTION DEALS WITH GUI FUNCTIONALITY #########################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

    def change_excel_path(self):

        # Define logger
        self.logger.debug(f'Running function change_video_folder()')

        # Get new Excel path
        self.annotation_file_path = utils.get_excel_path(self.annotation_file_path, 0, self.video_folder_path, self.crop_mode)

        # Reload the window with the new input
        self.reload(False, True)

    def change_video_folder(self):
        self.logger.debug(f'Running function change_video_folder()')

        # Set loaded to false as when video folder is changed, the GUI must be reloaded.
        self.loaded = False

        # Get new video fodler and reload the GUI
        self.video_folder_path, self.scanned_folders, self.dir_hierarchy = utils.get_video_folder(self.video_folder_path, 0)
        self.reload(False, True)

    def switch_folder(self, which):

        # Define logger
        self.logger.debug(f'Running function switch_folder({which})')

        # Set loaded to false as when video folder is changed, the GUI must be reloaded.
        self.loaded = False

        # Check if there is another folder either left or right in the directory hierarchy and if yes, then switch and reload.
        index = self.scanned_folders.index(os.path.basename(os.path.normpath(self.video_folder_path)))
        if index > 0 and which == "left":
            self.video_folder_path = os.path.join(os.path.dirname(self.video_folder_path), self.scanned_folders[index - 1])
            self.reload(False, True)
        if (index + 1) < len(self.scanned_folders) and which == "right":
            self.video_folder_path = os.path.join(os.path.dirname(self.video_folder_path), self.scanned_folders[index + 1])
            self.reload(False, True)

    def set_ocr_roi(self, video_filepath):

        # function that will open a frame with an image and prompt the user to drag a rectangle around the text and the
        # top left and bottom right coordinates will be saved in the settings_crop.ini file
        global x_coordinate
        global y_coordinate
        global width
        global height
        self.logger.debug(f'Running function set_ocr_roi({video_filepath})')
        def draw_rectangle(event, x, y, flags, param):
            global x_coordinate
            global y_coordinate
            global width
            global height
            global cap
            global text_roi
            frame = self.cap.read()[1]
            if event == cv2.EVENT_LBUTTONDOWN:
                self.x_coordinate = x
                self.y_coordinate = y
            elif event == cv2.EVENT_LBUTTONUP:
                self.width = x - self.x_coordinate
                self.height = y - self.y_coordinate
                self.text_roi = (self.x_coordinate, self.y_coordinate, self.width, self.height)
                cv2.rectangle(frame, (self.x_coordinate, self.y_coordinate), (x, y), (0, 255, 0), 2)
                cv2.imshow('image', frame)
                # cv2.waitKey(0)

        # Read settings from settings_crop.ini
        self.config.read('settings_crop.ini', encoding='utf-8')
        try:
            self.x_coordinate = int(self.config['OCR settings'].get('x_coordinate', '0').strip())
            self.y_coordinate = int(self.config['OCR settings'].get('y_coordinate', '0').strip())
            self.width = int(self.config['OCR settings'].get('width', '500').strip())
            self.height = int(self.config['OCR settings'].get('height', '40').strip())
        except ValueError:
            # Handle cases where conversion to integer fails
            logger.info('Error: Invalid integer value found in settings_crop.ini')
        # check if video_filepath is valid path to a video file
        if not os.path.isfile(video_filepath) or not video_filepath.endswith(".mp4"):
            logger.info('Error: Invalid video file path')
            return

        self.cap = cv2.VideoCapture(video_filepath)
        # Create a window and pass it to the mouse callback function
        cv2.namedWindow('image')
        # Make the window topmost
        cv2.setWindowProperty('image', cv2.WND_PROP_TOPMOST, 1)
        # display rectangle on image from the text_roi coordinates
        cv2.setMouseCallback('image', draw_rectangle)
        while (self.cap.isOpened()):
            ret, frame = self.cap.read()
            if ret:
                cv2.rectangle(frame, (self.x_coordinate, self.y_coordinate), (self.x_coordinate + self.width, self.y_coordinate + self.height),
                              (0, 255, 0),
                              2)
                cv2.imshow('image', frame)
                k = cv2.waitKey(1) & 0xFF
                if k == 27:
                    cv2.destroyAllWindows()
                    break
            else:
                break
        self.cap.release()
        cv2.destroyAllWindows()
        # Save settings to settings_crop.ini
        self.config['OCR settings']['x_coordinate'] = str(self.x_coordinate)
        self.config['OCR settings']['y_coordinate'] = str(self.y_coordinate)
        self.config['OCR settings']['width'] = str(self.width)
        self.config['OCR settings']['height'] = str(self.height)
        with open('settings_crop.ini', 'w', encoding='utf-8') as configfile:
            self.config.write(configfile)

    def update_entries(self, index, original_points):
        global points_of_interest_entry
        global modified_frames
        global logger
        logger.debug(f"Running function update_entries({index}, {original_points})")
        for each in range(index + 1, len(points_of_interest_entry)):
            if len(points_of_interest_entry[each]) == 0:
                points_of_interest_entry[each] = points_of_interest_entry[index].copy()
            else:
                if points_of_interest_entry[each] == original_points:
                    points_of_interest_entry[each] = points_of_interest_entry[index].copy()
                else:
                    break
        first = 1
        for each in range(index, len(video_filepaths)):
            first = first - 1
            if first >= 0 and (
                    index == 0 or not points_of_interest_entry[max(index - 1, 0)] == points_of_interest_entry[index]):
                modified_frames.append(each)
            update_button_image(frames[each].copy(), (max(each, 0) // 6), (each - ((max(each, 0) // 6) * 6)), first)

    def get_mouse_position(event, x, y, flags, mode, i, j):
        global points_of_interest_entry
        global modified_frames
        if event == cv2.EVENT_LBUTTONUP:
            index = j + ((i) * 6)
            if flags & cv2.EVENT_FLAG_SHIFTKEY:
                closest_point = None
                closest_distance = float('inf')
                for point in points_of_interest_entry[index]:
                    distance = math.sqrt((x - point[0]) ** 2 + (y - point[1]) ** 2)
                    if distance < closest_distance:
                        closest_point = point
                        closest_distance = distance
                if closest_distance < 30:
                    points_of_interest_entry[index].remove(closest_point)
                print(f"Flow: retrieved POIs: {points_of_interest_entry}")
                cv2.destroyAllWindows()
            elif flags & cv2.EVENT_FLAG_CTRLKEY and flags & cv2.EVENT_FLAG_ALTKEY:
                points_of_interest_entry[index] = []
                if index in modified_frames:
                    modified_frames.remove(index)
                cv2.destroyAllWindows()
            else:
                if not mode == 0:
                    points_of_interest_entry[index].append((x, y))
                else:
                    points_of_interest_entry.append((x, y))
                print(f"Flow: retrieved POIs: {points_of_interest_entry}")
                cv2.destroyAllWindows()

    def update_button_image(frame, i, j, first):
        global points_of_interest_entry
        global modified_frames
        global button_images
        global buttons
        global logger
        logger.debug(f"Running function update_button_image({i}, {j}, {first})")
        index = j + ((i) * 6)
        frame = frame.copy()
        height, width, channels = frame.shape
        for point in points_of_interest_entry[index]:
            x1, y1 = max(0, point[0] - (crop_size // 2)), max(0, point[1] - (crop_size // 2))
            x2, y2 = min(width, point[0] + (crop_size // 2)), min(height, point[1] + (crop_size // 2))
            cv2.rectangle(frame, (point[0] - 30, point[1] - 30), (point[0] + 30, point[1] + 30), (0, 255, 0), 2)
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 3)
            cv2.line(frame, (point[0], point[1]), (x1, y1), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            cv2.line(frame, (point[0], point[1]), (x1, y2), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            cv2.line(frame, (point[0], point[1]), (x2, y1), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            cv2.line(frame, (point[0], point[1]), (x2, y2), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            pos_off = 0
            rectangles = [
                [(
                 max(pos_off, min(((point[0] - crop_size // 2) + offset_range) - pos_off, width - crop_size - pos_off)),
                 max(pos_off,
                     min(((point[1] - crop_size // 2) + offset_range) - pos_off, height - crop_size - pos_off))),
                 (max(crop_size - pos_off,
                      min(((point[0] + crop_size // 2) + offset_range) - pos_off, width - pos_off)),
                  max(crop_size - pos_off,
                      min(((point[1] + crop_size // 2) + offset_range) - pos_off, height - pos_off)))],
                # Rectangle 1: (upper_left, bottom_right)
                [(max(pos_off,
                      min(((point[0] - crop_size // 2) - offset_range) - pos_off, width - crop_size - pos_off)),
                  max(pos_off,
                      min(((point[1] - crop_size // 2) - offset_range) - pos_off, height - crop_size - pos_off))),
                 (max(crop_size - pos_off,
                      min(((point[0] + crop_size // 2) - offset_range) - pos_off, width - pos_off)),
                  max(crop_size - pos_off,
                      min(((point[1] + crop_size // 2) - offset_range) - pos_off, height - pos_off)))],
                # Rectangle 2: (upper_left, bottom_right)
                [(max(pos_off,
                      min(((point[0] - crop_size // 2) + offset_range) - pos_off, width - crop_size - pos_off)),
                  max(pos_off,
                      min(((point[1] - crop_size // 2) - offset_range) - pos_off, height - crop_size - pos_off))),
                 (max(crop_size - pos_off,
                      min(((point[0] + crop_size // 2) + offset_range) - pos_off, width - pos_off)),
                  max(crop_size - pos_off,
                      min(((point[1] + crop_size // 2) - offset_range) - pos_off, height - pos_off)))],
                # Rectangle 3: (upper_left, bottom_right)
                [(max(pos_off,
                      min(((point[0] - crop_size // 2) - offset_range) - pos_off, width - crop_size - pos_off)),
                  max(pos_off,
                      min(((point[1] - crop_size // 2) + offset_range) - pos_off, height - crop_size - pos_off))),
                 (max(crop_size - pos_off,
                      min(((point[0] + crop_size // 2) - offset_range) - pos_off, width - pos_off)),
                  max(crop_size - pos_off,
                      min(((point[1] + crop_size // 2) + offset_range) - pos_off, height - pos_off)))]
                # Rectangle 4: (upper_left, bottom_right)
            ]

            # Find the overlapping area
            x_min = max(rect[0][0] for rect in rectangles)
            y_min = max(rect[0][1] for rect in rectangles)
            x_max = min(rect[1][0] for rect in rectangles)
            y_max = min(rect[1][1] for rect in rectangles)

            # Calculate the coordinates of the overlapping area
            cv2.rectangle(frame, (x_min, y_min), (x_max, y_max), (0, 255, 0), 2)

        if (first >= 0 or index in modified_frames) and (
                (index == 0 and not len(points_of_interest_entry[0]) == 0) or not points_of_interest_entry[
                                                                                      max(index - 1, 0)] ==
                                                                                  points_of_interest_entry[index]):
            # Define the ROI
            frame = cv2.resize(frame, (276, 156), interpolation=cv2.INTER_AREA)
            height, width, channels = frame.shape
            x, y, w, h = 0, 0, height // 5, height // 5
            roi = frame[y:y + h, x:x + w]

            # Load the overlay image
            overlay = cv2.imread('resources/img/sp.png', cv2.IMREAD_UNCHANGED)

            # Resize the overlay image to fit the ROI
            overlay = cv2.resize(overlay, (w, h))

            # Merge the overlay image and the ROI
            alpha = overlay[:, :, 3] / 255.0
            alpha = cv2.merge([alpha, alpha, alpha])
            background = cv2.multiply(1.0 - alpha, roi, dtype=cv2.CV_8UC3)
            foreground = cv2.multiply(alpha, overlay[:, :, 0:3], dtype=cv2.CV_8UC3)
            result = cv2.add(background, foreground)

            # Copy the result back into the original image
            frame[y:y + h, x:x + w] = result
        pil_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(pil_img, mode='RGB')
        # Resize image to fit button
        pil_img = pil_img.resize((276, 156))
        # Convert PIL image to tkinter image
        img = ImageTk.PhotoImage(pil_img)
        button_images[index] = img
        buttons[i][j].configure(image=button_images[index])

    def on_key_press(event):
        global b_l
        global b_r
        global u_l
        global u_r
        print("execd")
        if keyboard.is_pressed('left') and keyboard.is_pressed('down'):
            b_l = b_l * -1
        elif keyboard.is_pressed('right') and keyboard.is_pressed('down'):
            b_r = b_r * -1
        elif keyboard.is_pressed('left') and keyboard.is_pressed('up'):
            u_l = u_l * -1
        elif keyboard.is_pressed('right') and keyboard.is_pressed('up'):
            u_r = u_r * -1
        elif keyboard.is_pressed('q'):
            b_l = b_l * -1
            b_r = b_r * -1
            u_l = u_l * -1
            u_r = u_r * -1

    def on_button_click(i, j, button_images):
        global points_of_interest_entry
        global logger
        global b_l
        global b_r
        global u_l
        global u_r
        logger.debug(f"Running function on_button_click({i}, {j})")
        index = j + ((i) * 6)
        mode = 1
        frame_tmp = frames[index].copy()
        b_l = -1
        b_r = -1
        u_l = -1
        u_r = -1
        # Ask the user if they want to select additional points of interest
        while True:
            original_points = points_of_interest_entry[index].copy()
            frame = frame_tmp.copy()

            # Draw a rectangle around the already selected points of interest
            height, width, channels = frame.shape
            for point in points_of_interest_entry[index]:
                x1, y1 = max(0, point[0] - (crop_size // 2)), max(0, point[1] - (crop_size // 2))
                x2, y2 = min(width, point[0] + (crop_size // 2)), min(height, point[1] + (crop_size // 2))
                cv2.rectangle(frame, (point[0] - 30, point[1] - 30), (point[0] + 30, point[1] + 30), (0, 255, 0), 2)
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 3)
                cv2.line(frame, (point[0], point[1]), (x1, y1), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
                cv2.line(frame, (point[0], point[1]), (x1, y2), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
                cv2.line(frame, (point[0], point[1]), (x2, y1), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
                cv2.line(frame, (point[0], point[1]), (x2, y2), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
                if True:
                    if b_r > 0:
                        pos_off = 0
                        o_x1 = max(pos_off,
                                   min(((point[0] - crop_size // 2) + offset_range) - pos_off,
                                       width - crop_size - pos_off))
                        o_y1 = max(pos_off,
                                   min(((point[1] - crop_size // 2) + offset_range) - pos_off,
                                       height - crop_size - pos_off))
                        o_x2 = max(crop_size - pos_off,
                                   min(((point[0] + crop_size // 2) + offset_range) - pos_off, width - pos_off))
                        o_y2 = max(crop_size - pos_off,
                                   min(((point[1] + crop_size // 2) + offset_range) - pos_off, height - pos_off))
                        cv2.rectangle(frame, (o_x1, o_y1), (o_x2, o_y2), (255, 229, 0), 2)

                        # Add label text
                        label_text = 'BR'
                        label_color = (255, 229, 0)
                        label_font = cv2.FONT_HERSHEY_SIMPLEX
                        label_scale = 1
                        label_thickness = 2

                        (label_width, label_height), _ = cv2.getTextSize(label_text, label_font, label_scale,
                                                                         label_thickness)
                        label_x = o_x1 + (o_x2 - o_x1) // 2 - label_width // 2
                        label_y = o_y1 + (o_y2 - o_y1) // 2 + label_height // 2

                        # cv2.putText(frame, label_text, (label_x, label_y), label_font, label_scale, label_color,
                        #             label_thickness)
                        # cv2.line(frame, (o_x1, o_y1),
                        #          (label_x - (label_width // 4), label_y - label_height - (label_height // 4)),
                        #          label_color, label_thickness)
                    if u_l > 0:
                        pos_off = 0
                        o_x1 = max(pos_off,
                                   min(((point[0] - crop_size // 2) - offset_range) - pos_off,
                                       width - crop_size - pos_off))
                        o_y1 = max(pos_off,
                                   min(((point[1] - crop_size // 2) - offset_range) - pos_off,
                                       height - crop_size - pos_off))
                        o_x2 = max(crop_size - pos_off,
                                   min(((point[0] + crop_size // 2) - offset_range) - pos_off, width - pos_off))
                        o_y2 = max(crop_size - pos_off,
                                   min(((point[1] + crop_size // 2) - offset_range) - pos_off, height - pos_off))
                        cv2.rectangle(frame, (o_x1, o_y1), (o_x2, o_y2), (255, 229, 0), 2)

                        # Add label text
                        label_text = 'UL'
                        label_color = (255, 229, 0)
                        label_font = cv2.FONT_HERSHEY_SIMPLEX
                        label_scale = 1
                        label_thickness = 2

                        (label_width, label_height), _ = cv2.getTextSize(label_text, label_font, label_scale,
                                                                         label_thickness)
                        label_x = o_x1 + (o_x2 - o_x1) // 2 - label_width // 2
                        label_y = o_y1 + (o_y2 - o_y1) // 2 + label_height // 2

                        # cv2.putText(frame, label_text, (label_x, label_y), label_font, label_scale, label_color,
                        #             label_thickness)
                        # cv2.line(frame, (o_x1, o_y1),
                        #          (label_x - (label_width // 4), label_y - label_height - (label_height // 4)),
                        #          label_color, label_thickness)
                    if u_r > 0:
                        pos_off = 0
                        o_x1 = max(pos_off,
                                   min(((point[0] - crop_size // 2) + offset_range) - pos_off,
                                       width - crop_size - pos_off))
                        o_y1 = max(pos_off,
                                   min(((point[1] - crop_size // 2) - offset_range) - pos_off,
                                       height - crop_size - pos_off))
                        o_x2 = max(crop_size - pos_off,
                                   min(((point[0] + crop_size // 2) + offset_range) - pos_off, width - pos_off))
                        o_y2 = max(crop_size - pos_off,
                                   min(((point[1] + crop_size // 2) - offset_range) - pos_off, height - pos_off))
                        cv2.rectangle(frame, (o_x1, o_y1), (o_x2, o_y2), (255, 229, 0), 2)

                        # Add label text
                        label_text = 'UR'
                        label_color = (255, 229, 0)
                        label_font = cv2.FONT_HERSHEY_SIMPLEX
                        label_scale = 1
                        label_thickness = 2

                        (label_width, label_height), _ = cv2.getTextSize(label_text, label_font, label_scale,
                                                                         label_thickness)
                        label_x = o_x1 + (o_x2 - o_x1) // 2 - label_width // 2
                        label_y = o_y1 + (o_y2 - o_y1) // 2 + label_height // 2

                        # cv2.putText(frame, label_text, (label_x, label_y), label_font, label_scale, label_color,
                        #             label_thickness)
                        # cv2.line(frame, (o_x1, o_y1),
                        #          (label_x - (label_width // 4), label_y - label_height - (label_height // 4)),
                        #          label_color, label_thickness)
                    if b_l > 0:
                        pos_off = 0
                        o_x1 = max(pos_off,
                                   min(((point[0] - crop_size // 2) - offset_range) - pos_off,
                                       width - crop_size - pos_off))
                        o_y1 = max(pos_off,
                                   min(((point[1] - crop_size // 2) + offset_range) - pos_off,
                                       height - crop_size - pos_off))
                        o_x2 = max(crop_size - pos_off,
                                   min(((point[0] + crop_size // 2) - offset_range) - pos_off, width - pos_off))
                        o_y2 = max(crop_size - pos_off,
                                   min(((point[1] + crop_size // 2) + offset_range) - pos_off, height - pos_off))
                        cv2.rectangle(frame, (o_x1, o_y1), (o_x2, o_y2), (255, 229, 0), 2)

                        # Add label text
                        label_text = 'BL'
                        label_color = (255, 229, 0)
                        label_font = cv2.FONT_HERSHEY_SIMPLEX
                        label_scale = 1
                        label_thickness = 2

                        (label_width, label_height), _ = cv2.getTextSize(label_text, label_font, label_scale,
                                                                         label_thickness)
                        label_x = o_x1 + (o_x2 - o_x1) // 2 - label_width // 2
                        label_y = o_y1 + (o_y2 - o_y1) // 2 + label_height // 2

                        # cv2.putText(frame, label_text, (label_x, label_y), label_font, label_scale, label_color,
                        #             label_thickness)
                        # cv2.line(frame, (o_x1, o_y1),
                        #          (label_x - (label_width // 4), label_y - label_height - (label_height // 4)),
                        #          label_color, label_thickness)
                # Define rectangles by their upper left and bottom right corners
                pos_off = 0
                rectangles = [
                    [(max(pos_off,
                          min(((point[0] - crop_size // 2) + offset_range) - pos_off, width - crop_size - pos_off)),
                      max(pos_off,
                          min(((point[1] - crop_size // 2) + offset_range) - pos_off, height - crop_size - pos_off))),
                     (max(crop_size - pos_off,
                          min(((point[0] + crop_size // 2) + offset_range) - pos_off, width - pos_off)),
                      max(crop_size - pos_off,
                          min(((point[1] + crop_size // 2) + offset_range) - pos_off, height - pos_off)))],
                    # Rectangle 1: (upper_left, bottom_right)
                    [(max(pos_off,
                          min(((point[0] - crop_size // 2) - offset_range) - pos_off, width - crop_size - pos_off)),
                      max(pos_off,
                          min(((point[1] - crop_size // 2) - offset_range) - pos_off, height - crop_size - pos_off))),
                     (max(crop_size - pos_off,
                          min(((point[0] + crop_size // 2) - offset_range) - pos_off, width - pos_off)),
                      max(crop_size - pos_off,
                          min(((point[1] + crop_size // 2) - offset_range) - pos_off, height - pos_off)))],
                    # Rectangle 2: (upper_left, bottom_right)
                    [(max(pos_off,
                          min(((point[0] - crop_size // 2) + offset_range) - pos_off, width - crop_size - pos_off)),
                      max(pos_off,
                          min(((point[1] - crop_size // 2) - offset_range) - pos_off, height - crop_size - pos_off))),
                     (max(crop_size - pos_off,
                          min(((point[0] + crop_size // 2) + offset_range) - pos_off, width - pos_off)),
                      max(crop_size - pos_off,
                          min(((point[1] + crop_size // 2) - offset_range) - pos_off, height - pos_off)))],
                    # Rectangle 3: (upper_left, bottom_right)
                    [(max(pos_off,
                          min(((point[0] - crop_size // 2) - offset_range) - pos_off, width - crop_size - pos_off)),
                      max(pos_off,
                          min(((point[1] - crop_size // 2) + offset_range) - pos_off, height - crop_size - pos_off))),
                     (max(crop_size - pos_off,
                          min(((point[0] + crop_size // 2) - offset_range) - pos_off, width - pos_off)),
                      max(crop_size - pos_off,
                          min(((point[1] + crop_size // 2) + offset_range) - pos_off, height - pos_off)))]
                    # Rectangle 4: (upper_left, bottom_right)
                ]

                # Find the overlapping area
                x_min = max(rect[0][0] for rect in rectangles)
                y_min = max(rect[0][1] for rect in rectangles)
                x_max = min(rect[1][0] for rect in rectangles)
                y_max = min(rect[1][1] for rect in rectangles)

                # Calculate the coordinates of the overlapping area
                cv2.rectangle(frame, (x_min, y_min), (x_max, y_max), (0, 255, 0), 2)

            # Display the image with the rectangles marking the already selected points of interest
            cv2.imshow("Frame", frame)
            cv2.setWindowProperty("Frame", cv2.WND_PROP_TOPMOST, 1)
            screen_width, screen_height = cv2.getWindowImageRect("Frame")[2:]
            window_width, window_height = int(1 * screen_width), int(1 * screen_height)
            cv2.resizeWindow("Frame", window_width, window_height)
            cv2.moveWindow("Frame", int((screen_width // 2) - (window_width // 2)), 0)

            # Prompt the user to click on the next point of interest
            cv2.setMouseCallback("Frame",
                                 lambda event, x, y, flags, mode: get_mouse_position(event, x, y, flags, mode, i, j),
                                 mode)
            key = cv2.waitKey(0)
            if key == 27 or key == 13:  # Check if the Esc key was pressed
                cv2.destroyAllWindows()
                break
            elif key == ord("q"):
                u_l = u_l * -1
            elif key == ord("w"):
                u_r = u_r * -1
            elif key == ord("e"):
                b_l = b_l * -1
            elif key == ord("r"):
                b_r = b_r * -1
            update_entries(index, original_points)

    def update_crop_mode(var):
        global crop_mode
        global logger
        logger.debug(f"Running function update_crop_mode({var})")
        crop_mode = var

    def save_progress():
        global auto_processing
        global points_of_interest_entry
        global video_filepaths
        global logger
        logger.debug(f"Running function save_progress()")
        result = ask_yes_no("Do you want to save the settings? This will overwrite any previous saves.")
        if result:
            if check_path(video_folder_path, 0):
                # Create an in-memory file object
                filepath = os.path.join(video_folder_path, 'crop_information.pkl')
                with open(filepath, 'wb') as f:
                    # Use the pickle module to write the data to the file
                    pickle.dump([auto_processing.get(), points_of_interest_entry, video_filepaths], f)
            else:
                # display message box with error message
                messagebox.showerror("Error", "Invalid video folder path")

    def load_progress():
        global auto_processing
        global points_of_interest_entry
        global video_filepaths
        global frames
        global loaded
        global auto
        global root
        global logger
        logger.debug(f"Running function load_progress()")
        result = ask_yes_no("Do you want to load settings? This will overwrite any unsaved progress.")
        if result:
            if check_path(video_folder_path, 0):
                # Create an in-memory file object
                filepath = os.path.join(video_folder_path, 'crop_information.pkl')
                if os.path.isfile(filepath):
                    try:
                        if root.winfo_exists():
                            root.destroy()
                    except:
                        print("Error: Unexpected, window destroyed before reference. Odpruženo.")
                    points_of_interest_entry = []
                    with open(filepath, 'rb') as f:
                        data = pickle.load(f)
                    # print(data)
                    auto = data[0]
                    points_of_interest_entry = data[1].copy()
                    video_filepaths_new = data[2].copy()
                    set1 = set(video_filepaths_new)
                    set2 = set(video_filepaths)
                    if not set1 == set2:
                        messagebox.showinfo("Discrepancy detected",
                                            "The contents of the video folder have changed since the save has been made. Cannot load the progress. Please start over.")
                        reload_points_of_interest()
                    else:
                        video_filepaths = []
                        video_filepaths = video_filepaths_new.copy()
                    reload(1, False)
                    # load_videos()
                    # load_video_frames()
                    # loaded = 1
                    # open_ICCS_window()
                else:
                    messagebox.showinfo("No save detected",
                                        "There are no save files in the current directory.")
            else:
                # display message box with error message
                messagebox.showerror("Error", "Invalid video folder path")