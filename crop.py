import pandas as pd
import os
import subprocess
import re
import cv2
import pytesseract
import configparser
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageTk
import pickle
import datetime
import sys
import random
import openpyxl
import math
import time
from hachoir.metadata import extractMetadata
from hachoir.parser import createParser
import logging
import xlwings as xw
import keyboard
from ultralytics import YOLO
import torch
import shutil
import asyncio
import cProfile
import tracemalloc
from utils import ask_yes_no
from utils import create_dir
from utils import scan_default_folders
from utils import check_path
from utils import get_video_folder
from utils import get_excel_path
tracemalloc.start()

def config_read():
    global ocr_tesseract_path
    global video_folder_path
    global annotation_file_path
    global output_folder
    global scan_folders
    global crop_mode
    global frame_skip
    global frames_per_visit
    global filter_visitors
    global yolo_processing
    global default_label_category
    global yolo_conf
    global randomize
    global whole_frame
    global cropped_frames
    global crop_size
    global offset_range
    global config
    global prefix
    global logger
    logger.debug('Running function config_read()')
    # load config or create the file
    # Set default values
    config = configparser.ConfigParser()
    config['Resource Paths'] = {
        'OCR_tesseract_path': 'C:/Program Files/Tesseract-OCR/tesseract.exe',
        'video_folder_path': '',
        'annotation_file_path': '',
        'output_folder': 'output'
    }
    config['OCR settings'] = {
        'x_coordinate': '0',
        'y_coordinate': '0',
        'width': '500',
        'height': '40'
    }
    config['Crop settings'] = {
        'crop_mode': '1',
        'crop_interval_frames': '30',
        'frames_per_visit': '0',
        'filter_visitors': '0',
        'yolo_processing': '0',
        'default_label_category': '6',
        'yolo_conf': '0.25',
        'randomize_interval': '0',
        'export_whole_frame': '0',
        'export_crops': '1',
        'crop_size': '640',
        'random_offset_range': '600',
        'filename_prefix': ''
    }
    config['Workflow settings'] = {
        'Scan_default_folders': '1'
    }

    # Check if settings_crop.ini exists, and create it with default values if not
    if not os.path.exists('settings_crop.ini'):
        with open('settings_crop.ini', 'w', encoding='utf-8') as configfile:
            config.write(configfile)

    # Read settings from settings_crop.ini
    config.read('settings_crop.ini', encoding='utf-8')

    # Get values from the config file
    try:
        ocr_tesseract_path = config['Resource Paths'].get('OCR_tesseract_path',
                                                          'C:/Program Files/Tesseract-OCR/tesseract.exe').strip()
        video_folder_path = config['Resource Paths'].get('video_folder_path', '').strip()
        annotation_file_path = config['Resource Paths'].get('annotation_file_path', '').strip()
        output_folder = config['Resource Paths'].get('output_folder', 'output').strip()
        scan_folders = config['Workflow settings'].get('Scan_default_folders', '0').strip()
    except ValueError:
        print('Error: Invalid folder/file path found in settings_crop.ini')
    # Get crop values from config
    try:
        crop_mode = int(config['Crop settings'].get('crop_mode', '1').strip())
        frame_skip = int(config['Crop settings'].get('crop_interval_frames', '30').strip())
        frames_per_visit = int(config['Crop settings'].get('frames_per_visit', '0').strip())
        filter_visitors = int(config['Crop settings'].get('filter_visitors', '0').strip())
        yolo_processing = int(config['Crop settings'].get('yolo_processing', '0').strip())
        default_label_category = int(config['Crop settings'].get('default_label_category', '6').strip())
        yolo_conf = float(config['Crop settings'].get('yolo_conf', '0.25').strip())
        randomize = int(config['Crop settings'].get('randomize_interval', '0').strip())
        whole_frame = int(config['Crop settings'].get('export_whole_frame', '0').strip())
        cropped_frames = int(config['Crop settings'].get('export_crops', '1').strip())
        crop_size = int(config['Crop settings'].get('crop_size', '640').strip())
        offset_range = int(config['Crop settings'].get('random_offset_range', '600').strip())
        prefix = config['Crop settings'].get('filename_prefix', '').strip()
    except ValueError:
        print('Error: Invalid crop settings specified in settings_crop.ini')

def reload_points_of_interest():
    global logger
    logger.debug('Running function reload_points_of_interest()')
    global points_of_interest_entry
    points_of_interest_entry = []
    for e in range(len(video_filepaths)):
        points_of_interest_entry.append([])

def change_excel_path():
    global logger
    global annotation_file_path
    global root
    logger.debug(f'Running function change_video_folder()')
    annotation_file_path = get_excel_path(annotation_file_path, 0, video_folder_path, crop_mode)
    reload(0, True)

def change_video_folder():
    global video_folder_path
    global points_of_interest_entry
    global scaned_folders
    global tree_allow
    global loaded
    global root
    global logger
    logger.debug(f'Running function change_video_folder()')
    loaded = 0
    video_folder_path, scaned_folders, tree_allow = get_video_folder(video_folder_path, 0)
    reload(0, True)

def switch_folder(which):
    global scaned_folders
    global video_folder_path
    global loaded
    global logger
    logger.debug(f'Running function switch_folder({which})')
    loaded = 0
    index = scaned_folders.index(os.path.basename(os.path.normpath(video_folder_path)))
    if index > 0 and which == "left":
        video_folder_path = os.path.join(os.path.dirname(video_folder_path), scaned_folders[index - 1])
        reload(0, True)
    if (index + 1) < len(scaned_folders) and which == "right":
        video_folder_path = os.path.join(os.path.dirname(video_folder_path), scaned_folders[index + 1])
        reload(0, True)


def load_videos():
    global video_filepaths
    global logger
    logger.debug('Running function load_videos()')
    # Check if the video folder path is valid
    if check_path(video_folder_path, 0):
        # Load videos
        video_filepaths = []
        video_filepaths = [os.path.join(video_folder_path, f) for f in os.listdir(video_folder_path) if f.endswith('.mp4')]
    else:
        messagebox.showerror("Error", "Invalid video folder path")
        video_filepaths = []

def set_ocr_roi(video_filepath):

    # function that will open a frame with an image and prompt the user to drag a rectangle around the text and the
    # top left and bottom right coordinates will be saved in the settings_crop.ini file
    global x_coordinate
    global y_coordinate
    global width
    global height
    global cap
    global config
    global logger
    logger.debug(f'Running function set_ocr_roi({video_filepath})')

    def draw_rectangle(event, x, y, flags, param):
        global x_coordinate
        global y_coordinate
        global width
        global height
        global cap
        global text_roi
        frame = cap.read()[1]
        if event == cv2.EVENT_LBUTTONDOWN:
            x_coordinate = x
            y_coordinate = y
        elif event == cv2.EVENT_LBUTTONUP:
            width = x - x_coordinate
            height = y - y_coordinate
            text_roi = (x_coordinate, y_coordinate, width, height)
            cv2.rectangle(frame, (x_coordinate, y_coordinate), (x, y), (0, 255, 0), 2)
            cv2.imshow('image', frame)
            # cv2.waitKey(0)

    # Read settings from settings_crop.ini
    config.read('settings_crop.ini', encoding='utf-8')
    try:
        x_coordinate = int(config['OCR settings'].get('x_coordinate', '0').strip())
        y_coordinate = int(config['OCR settings'].get('y_coordinate', '0').strip())
        width = int(config['OCR settings'].get('width', '500').strip())
        height = int(config['OCR settings'].get('height', '40').strip())
    except ValueError:
        # Handle cases where conversion to integer fails
        print('Error: Invalid integer value found in settings_crop.ini')
    # check if video_filepath is valid path to a video file
    if not os.path.isfile(video_filepath) or not video_filepath.endswith(".mp4"):
        print('Error: Invalid video file path')
        return
    cap = cv2.VideoCapture(video_filepath)
    # Create a window and pass it to the mouse callback function
    cv2.namedWindow('image')
    # Make the window topmost
    cv2.setWindowProperty('image', cv2.WND_PROP_TOPMOST, 1)
    # display rectangle on image from the text_roi coordinates
    cv2.setMouseCallback('image', draw_rectangle)
    while (cap.isOpened()):
        ret, frame = cap.read()
        if ret:
            cv2.rectangle(frame, (x_coordinate, y_coordinate), (x_coordinate + width, y_coordinate + height),
                          (0, 255, 0),
                          2)
            cv2.imshow('image', frame)
            k = cv2.waitKey(1) & 0xFF
            if k == 27:
                cv2.destroyAllWindows()
                break
        else:
            break
    cap.release()
    cv2.destroyAllWindows()
    # Save settings to settings_crop.ini
    config['OCR settings']['x_coordinate'] = str(x_coordinate)
    config['OCR settings']['y_coordinate'] = str(y_coordinate)
    config['OCR settings']['width'] = str(width)
    config['OCR settings']['height'] = str(height)
    with open('settings_crop.ini', 'w', encoding='utf-8') as configfile:
        config.write(configfile)

def get_text_from_video(video_filepath, start_or_end):
    global x_coordinate
    global y_coordinate
    global width
    global height
    global config
    global cap
    global logger
    logger.debug(f'Running function get_text_from_video({video_filepath}, {start_or_end})')

    # Read settings from settings_crop.ini
    config.read('settings_crop.ini', encoding='utf-8')
    try:
        x_coordinate = int(config['OCR settings'].get('x_coordinate', '0').strip())
        y_coordinate = int(config['OCR settings'].get('y_coordinate', '0').strip())
        width = int(config['OCR settings'].get('width', '500').strip())
        height = int(config['OCR settings'].get('height', '40').strip())
    except ValueError:
        # Handle cases where conversion to integer fails
        print('Error: Invalid integer value found in settings_crop.ini')

    # Get the video capture and ROI defined
    cap = cv2.VideoCapture(video_filepath)
    text_roi = (x_coordinate, y_coordinate, width, height)  # x, y, width, height

    # Define which frame to scan - start of end?
    if start_or_end == "end":
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        second_to_last_frame_idx = total_frames - 5
        cap.set(cv2.CAP_PROP_POS_FRAMES, second_to_last_frame_idx)
    else:
        cap.set(cv2.CAP_PROP_POS_FRAMES, 24)

    # Get the video frame and process the image
    ret, frame = cap.read()
    if ret:
        #Crop the image and pre-process it
        #height, width, channels = frame.shape
        x, y, w, h = text_roi
        text_frame = frame[y:y + h, x:x + w]
        HSV_img = cv2.cvtColor(text_frame, cv2.COLOR_BGR2HSV)
        h, s, v = cv2.split(HSV_img)
        v = cv2.GaussianBlur(v, (1, 1), 0)
        thresh = cv2.threshold(v, 220, 255, cv2.THRESH_BINARY_INV)[1] #change the second number to change the threshold - anything over that value will be turned into white
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, ksize=(1, 1))
        thresh = cv2.dilate(thresh, kernel)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, ksize=(1, 1))
        thresh = cv2.erode(thresh, kernel)
        # text recognition
        OCR_text = pytesseract.image_to_string(thresh)

        # debug OCR file creation
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, "OCR images")
        create_dir(output_dir)
        OCR_file_name = ''.join([os.path.basename(video_filepath)[:-4], "_", start_or_end, ".png"])
        OCR_file_path = os.path.join(output_dir, OCR_file_name)
        #cv2.imwrite(OCR_file_path, thresh)
        with open(OCR_file_path, 'wb') as f:
            f.write(cv2.imencode('.png', thresh)[1].tobytes())
    else:
        OCR_text = "none"
    cap.release()
    return OCR_text, frame


def get_text_manually(frame):
    global sec_OCR
    global root
    global dialog
    global x_coordinate
    global y_coordinate
    global width
    global height
    global logger

    def submit_time(input_field):
        global dialog
        global sec_OCR
        global root
        global logger
        logger.debug(f'Running function submit_time({input_field})')
        text = input_field.get()
        if not text.isdigit() or len(text) != 2 or int(text) > 59:
            # execute code here for when text is not in "SS" format
            print(
                "Error: OCR detected text does not follow the expected format. Manual input is not in the correct format. The value will be set to an arbitrary 00.")
            text = '00'
        else:
            print("Error: OCR detected text does not follow the expected format. Resolved manually.")
        sec_OCR = text
        while True:
            try:
                if dialog.winfo_exists():
                    dialog.quit()
                    dialog.destroy()
                    break
            except:
                time.sleep(0.1)
                break

    # loop until the tkinter root window is created
    while True:
        try:
            if root.winfo_exists():
                break
        except:
            time.sleep(0.1)

    # Create the dialog window
    dialog = tk.Toplevel(root)
    try:
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
    except:
        screen_width = 1920
        screen_height = 1080

    dialog.wm_attributes("-topmost", 1)
    dialog.title("Time Input")

    # convert frame to tkinter image
    text_roi = (x_coordinate, y_coordinate, width, height)
    x, y, w, h = text_roi
    img_frame_width = min(screen_width // 2, w * 2)
    img_frame_height = min(screen_height // 2, h * 2)
    frame = frame[y:y + h, x:x + w]
    img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = cv2.resize(img, (img_frame_width, img_frame_height), Image.LANCZOS)
    img = Image.fromarray(img)
    img = ImageTk.PhotoImage(img)
    img_width = img.width()
    img_height = img.height()

    # Add image frame containing the video frame
    img_frame = tk.Frame(dialog, width=img_width, height=img_height)
    img_frame.pack(side=tk.TOP, pady=(0, 0))
    img_frame.pack_propagate(0)

    # Add image to image frame
    img_label = tk.Label(img_frame, image=img)
    img_label.pack(side=tk.TOP, pady=(0, 0))

    # Add label
    text_field = tk.Text(dialog, height=2, width=120, font=("Arial", 10))
    text_field.insert(tk.END,
                      "The OCR detection apparently failed.\nEnter the last two digits of the security camera watermark (number of seconds).\nThis will ensure cropping will happen at the right times")
    text_field.configure(state="disabled", highlightthickness=1, relief="flat", background=dialog.cget('bg'))
    text_field.tag_configure("center", justify="center")
    text_field.tag_add("center", "1.0", "end")
    text_field.pack(side=tk.TOP, padx=(0, 0))
    label = tk.Label(dialog, text="Enter text", font=("Arial", 10), background=dialog.cget('bg'))
    label.pack(pady=2)

    # Add input field
    j = 0
    input_field = tk.Entry(dialog, font=("Arial", 10), width=4)
    input_field.pack(pady=2)
    input_field.bind("<Return>", lambda j=j: submit_time(input_field))
    input_field.focus()

    # Add submit button
    submit_button = tk.Button(dialog, text="Submit", font=("Arial", 10),
                              command=lambda j=j: submit_time(input_field))
    submit_button.pack(pady=2)

    # Position the window
    dialog_width = dialog.winfo_reqwidth() + img_frame_width
    dialog_height = dialog.winfo_reqheight() + img_frame_height
    dialog_pos_x = int((screen_width // 2) - (img_frame_width // 2))
    dialog_pos_y = 0
    dialog.geometry(f"{dialog_width}x{dialog_height}+{dialog_pos_x}+{dialog_pos_y}")

    # Start the dialog
    dialog.mainloop()

    # When dialog is closed
    return_time = sec_OCR
    if len(return_time) > 0:
        success = True
    else:
        success = False
    return return_time, success

def get_metadata_from_video(video_filepath, start_or_end):
    if start_or_end == "start":
        try:
            parser = createParser(video_filepath)
            metadata = extractMetadata(parser)
            modify_date = str(metadata.get("creation_date"))
            return_time = modify_date[-2:]
            print("Flow: Obtained video start time from metadata.")
            success = True
        except:
            success = False
    elif start_or_end == "end":
        try:
            parser = createParser(video_filepath)
            metadata = extractMetadata(parser)
            modify_date = str(metadata.get("creation_date"))
            start_seconds = int(modify_date[-2:])
            duration = str(metadata.get("duration"))
            time_parts = duration.split(":")
            seconds = int(time_parts[2].split(".")[0])
            return_time = str(seconds + start_seconds)
            print("Flow: Obtained video end time from metadata.")
            success = True
        except:
            success = False
    return return_time, success

def process_OCR_text(detected_text, frame, video_filepath, start_or_end):
    global cap
    global sec_OCR
    global root
    global dialog
    global x_coordinate
    global y_coordinate
    global width
    global height
    global logger
    logger.debug(f'Running function process_OCR_text({detected_text}, {video_filepath}, {start_or_end})')
    return_time = "00"
    if "\n" in detected_text:
        detected_text = detected_text.replace("\n", "")
    if not len(detected_text) <= 23:
        detected_text = detected_text.rstrip()
        while not detected_text[-1].isdigit():
            detected_text = detected_text[:-1]
    correct_format = r"(0[0-9]|[1-5][0-9]):(0[0-9]|[1-5][0-9]):(0[0-9]|[1-5][0-9])"
    if re.match(correct_format, detected_text[-8:]):
        print(' '.join(["Flow:", "Text detection successful -", detected_text[-8:]]))
        return_time = detected_text[-2:]
        success = True
    else:
        print(' '.join(["Flow:", "Text detection failed -", detected_text]))
        success = False
    return return_time, success

# define function to get start and end times for each video file
def get_video_start_end_times(video_filepath):
    global logger
    logger.debug(f"Running function get_video_start_end_times({video_filepath})")
    video_filename = os.path.basename(video_filepath)
    print(' '.join(["Flow:", "Processing video file -", video_filepath]))

    # get start time
    # get the time from filename
    parts = video_filename[:-4].split("_")
    if len(parts) == 6:
        start_time_minutes = "_".join([parts[3], parts[4], parts[5]])
        print(' '.join(
            ["Flow: Video name format with prefixes detected. Extracted the time values -", start_time_minutes]))
    else:
        print(
            "Error: Some video file names have an unsupported format. Expected format is "
            "CO_LO1_SPPSPP1_YYYYMMDD_HH_MM. Script assumes format YYYYMMDD_HH_MM.")
        start_time_minutes = video_filename[:-4]
    start_time_seconds, success = get_metadata_from_video(video_filepath, "start")
    if not success:
        text, frame = get_text_from_video(video_filepath, "start")
        start_time_seconds, success = process_OCR_text(text, frame, video_filepath, "start")
        if not success:
            start_time_seconds, success = get_text_manually(frame)
    start_time_str = '_'.join([start_time_minutes, start_time_seconds])
    start_time = pd.to_datetime(start_time_str, format='%Y%m%d_%H_%M_%S')

    # get end time
    end_time_seconds, success = get_metadata_from_video(video_filepath, "end")
    if not success:
        text, frame = get_text_from_video(video_filepath, "end")
        end_time_seconds, success = process_OCR_text(text, frame, video_filepath, "end")
        if not success:
            end_time_seconds, success = get_text_manually(frame)
    try:
        parser = createParser(video_filepath)
        metadata = extractMetadata(parser)
        duration = str(metadata.get("duration"))
        time_parts = duration.split(":")
        delta = int(time_parts[1])
    except:
        delta = 15 + (int(end_time_seconds) // 60)
    #print(start_time_minutes)
    #print(end_time_seconds)
    end_time_seconds = str(int(end_time_seconds) % 60)
    end_time_str = pd.to_datetime('_'.join([start_time_minutes, end_time_seconds]), format='%Y%m%d_%H_%M_%S')
    end_time = end_time_str + pd.Timedelta(minutes=int(delta))
    return start_time, end_time

def evaluate_string_formula(cell):
    if isinstance(cell, (int, float)):
        # If the cell contains a number, return the value as is
        return cell
    elif cell.startswith('='):
        # If the cell contains an Excel formula, use openpyxl to evaluate it
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'].value = cell
        value = ws['A1'].value
        wb.close()
        return value
    else:
        # If the cell contains text, return the text as is
        return cell


def load_excel_table(file_path):
    global logger
    logger.debug(f"Running function load_excel_table({file_path})")
    # Define the columns to extract
    cols: list[int] = [0, 1, 2, 3, 4, 5, 15, 18, 19, 20, 21, 23, 24]

    # Read the Excel file, skipping the first two rows
    try:
        df = pd.read_excel(file_path, usecols=cols, skiprows=2, header=None,
                           converters={0: evaluate_string_formula, 1: evaluate_string_formula, 2: evaluate_string_formula,
                                       3: evaluate_string_formula, 4: evaluate_string_formula, 18: evaluate_string_formula})
    except ValueError as e:
        logger.error(f"Error reading Excel file {file_path}. Error message: {e}")
        # Open the Excel workbook using xlwings
        wb = xw.Book(file_path)

        sheet = wb.sheets[0]

        # Remove any filters
        if sheet.api.AutoFilterMode:
            sheet.api.AutoFilterMode = False

        #Save to temporary file
        create_dir("resources/exc")
        temp_file_path = os.path.join("resources/exc", "temp.xlsx")
        wb.save(temp_file_path)
        wb.close()

        #Read with pandas
        try:
            df = pd.read_excel(temp_file_path, usecols=cols, skiprows=2, header=None,
                           converters={0: evaluate_string_formula, 1: evaluate_string_formula,
                                       2: evaluate_string_formula,
                                       3: evaluate_string_formula, 4: evaluate_string_formula,
                                       18: evaluate_string_formula})
        except ValueError as e:
            logger.error(f"Attempted to fix errors in Excel file {file_path}. Attempt failed. Error message: {e}. Please fix the errors manually and try again.")
            return None
        logger.info(f"Attempted to remove filters from Excel file {file_path}. Saved a copy of the file to {temp_file_path}")

    print(f"Flow: Retrieved dataframe from Excel:\n{df}")
    filtered_df = df[df.iloc[:, 6] == 1]

    # Convert the month abbreviations in column 2 to month numbers
    months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
              'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
    filtered_df.iloc[:, 1] = filtered_df.iloc[:, 1].replace(months)
    filtered_df.iloc[:, 1] = filtered_df.iloc[:, 1].astype(int)
    filtered_df = filtered_df.copy()
    for i in range(5):
        j = i + 1
        filtered_df.iloc[:, j] = filtered_df.iloc[:, j].astype(int).apply(lambda x: f'{x:02}')
    filtered_df.loc[:, 13] = filtered_df.iloc[:, 11]
    filtered_df.iloc[:, 12] = filtered_df.iloc[:, 12]
    filtered_df.iloc[:, 11] = filtered_df.iloc[:, 0:6].apply(lambda x: f"{x[0]}{x[1]}{x[2]}_{x[3]}_{x[4]}_{x[5]}",
                                                            axis=1)
    filtered_df.iloc[:, 0] = filtered_df.iloc[:, 0].astype(int)
    print(f"Flow: Filtered dataframe:\n {filtered_df}")
    filtered_data = filtered_df.iloc[:, [7, 11]].values.tolist()

    # Get the column names
    column_1 = filtered_df.columns[12]
    column_2 = filtered_df.columns[13]

    # Count the number of NAs in each column
    na_count_1 = filtered_df[column_1].isna().sum()
    na_count_2 = filtered_df[column_2].isna().sum()

    # Check which column has fewer NAs
    if na_count_1 <= na_count_2:
        chosen_column = column_1
        other_column = column_2
    else:
        chosen_column = column_2
        other_column = column_1

    # Replace NAs in chosen_column with values from other_column if they are not NAs
    visitor_id = filtered_df[[chosen_column, other_column]].copy()
    visitor_id[chosen_column].fillna(visitor_id[other_column], inplace=True)
    visitor_id = visitor_id[[chosen_column]].values.tolist()
    print(visitor_id)

    annotation_data_array = filtered_data
    create_dir("resources/exc/")
    filtered_df.to_excel("resources/exc/output_filtered_crop.xlsx", index=False)
    return annotation_data_array, visitor_id

def load_csv(file_path):
    global logger
    logger.debug(f"Running function load_csv({file_path})")
    # Define the columns to extract
    cols: list[int] = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

    # Read the Excel file, skipping the first two rows - follow the custom format
    try:
        filtered_df = pd.read_excel(file_path, usecols=cols, skiprows=2, header=None,
                                    converters={0: evaluate_string_formula, 1: evaluate_string_formula,
                                                2: evaluate_string_formula,
                                                3: evaluate_string_formula, 4: evaluate_string_formula,
                                                6: evaluate_string_formula})
    except ValueError as e:
        logger.error(f"Error reading Excel file {file_path}. Error message: {e}")
        # Open the Excel workbook using xlwings
        wb = xw.Book(file_path)

        sheet = wb.sheets[0]

        # Remove any filters
        if sheet.api.AutoFilterMode:
            sheet.api.AutoFilterMode = False

        #Save to temporary file
        create_dir("resources/exc")
        temp_file_path = os.path.join("resources/exc", "temp.xlsx")
        wb.save(temp_file_path)
        wb.close()

        #Read with pandas
        try:
            filtered_df = pd.read_excel(file_path, usecols=cols, skiprows=2, header=None,
                                        converters={0: evaluate_string_formula, 1: evaluate_string_formula,
                                                    2: evaluate_string_formula,
                                                    3: evaluate_string_formula, 4: evaluate_string_formula,
                                                    6: evaluate_string_formula})
        except ValueError as e:
            logger.error(f"Attempted to fix errors in Excel file {file_path}. Attempt failed. Error message: {e}. Please fix the errors manually and try again.")
            return None
        logger.info(f"Attempted to remove filters from Excel file {file_path}. Saved a copy of the file to {temp_file_path}")
    print(f"Flow: Retrieved dataframe from Excel:\n{filtered_df}")

    # Convert the month abbreviations in column 2(1) to month numbers
    months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
              'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
    filtered_df.iloc[:, 1] = filtered_df.iloc[:, 1].replace(months)
    filtered_df.iloc[:, 1] = filtered_df.iloc[:, 1].astype(int)
    filtered_df = filtered_df.copy()
    for i in range(5):
        j = i + 1
        filtered_df.iloc[:, j] = filtered_df.iloc[:, j].astype(int).apply(lambda x: f'{x:02}')
    filtered_df.loc[:, 10] = filtered_df.iloc[:, 0:6].apply(lambda x: f"{x[0]}{x[1]}{x[2]}_{x[3]}_{x[4]}_{x[5]}",
                                                            axis=1)
    filtered_df.iloc[:, 0] = filtered_df.iloc[:, 0].astype(int)
    print(f"Flow: Filtered dataframe:\n {filtered_df}")

    # convert to list
    filtered_data = filtered_df.iloc[:, [6, 10]].values.tolist()

    annotation_data_array = filtered_data
    create_dir("resources/exc/")
    filtered_df.to_excel("resources/exc/output_filtered_crop.xlsx", index=False)
    return annotation_data_array


async def capture_crop(frame, point):
    global cap
    global fps
    global frame_number_start
    global visit_duration
    global crop_size
    global logger
    logger.debug(f"Running function capture_crop({point})")

    # Prepare local variables
    x, y = point

    # Add a random offset to the coordinates, but ensure they remain within the image bounds
    frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    # Check if any of the dimensions is smaller than crop_size and if so upscale the image to prevent crops smaller than desired crop_size
    if frame_height < crop_size or frame_width < crop_size:

        # Calculate the scaling factor to upscale the image
        scaling_factor = crop_size / min(frame_height, frame_width)

        # Calculate the new dimensions for the upscaled frame
        new_width = int(round(frame_width * scaling_factor))
        new_height = int(round(frame_height * scaling_factor))

        # Upscale the frame using cv2.resize with Lanczos upscaling algorithm
        frame = cv2.resize(frame, (new_width, new_height), interpolation=cv2.INTER_LANCZOS4)

    # Get the new frame size
    frame_height, frame_width = frame.shape[:2]

    # Calculate the coordinates for the area that will be cropped
    x_offset = random.randint(-offset_range, offset_range)
    y_offset = random.randint(-offset_range, offset_range)
    x1 = max(0, min(((x - crop_size // 2) + x_offset), frame_width - crop_size))
    y1 = max(0, min(((y - crop_size // 2) + y_offset), frame_height - crop_size))
    x2 = max(crop_size, min(((x + crop_size // 2) + x_offset), frame_width))
    y2 = max(crop_size, min(((y + crop_size // 2) + y_offset), frame_height))

    # Crop the image
    crop = frame[y1:y2, x1:x2]

    # Convert to correct color space
    crop_img = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
    if crop_img.shape[2] == 3:
        crop_img = cv2.cvtColor(crop_img, cv2.COLOR_BGR2RGB)

    # Return the cropped image and the coordinates for future reference
    return crop_img, x1, y1, x2, y2

async def generate_frames(frame, success, tag, index):
    global points_of_interest_entry
    global cap
    global fps
    global frame_number_start
    global visit_duration
    global logger
    global frame_skip
    global loop
    logger.debug(f"Running function generate_frames({index})")

    # Prepare name elements
    species = tag[-27:-19].replace("_", "")
    timestamp = tag[-18:-4]

    # Define local variables
    crop_counter = 1
    frame_skip_loc = frame_skip

    # Calculate the frame skip variable based on the limited number of frames per visit
    if frames_per_visit > 0:
        frame_skip_loc = int((visit_duration * fps)//frames_per_visit)
        if frame_skip_loc < 1:
            frame_skip_loc = 1

    # Loop through the video and crop y images every n-th frame
    frame_count = 0
    img_paths = []

    while success:
        # Crop images every n-th frame
        if int(frame_count % frame_skip_loc) == 0:
            for i, point in enumerate(points_of_interest_entry[index]):
                if cropped_frames == 1:
                    crop_img, x1, y1, x2, y2 = await capture_crop(frame, point)
                    img_path = f"./{output_folder}/{prefix}{species}_{timestamp}_{frame_number_start + frame_count}_{crop_counter}_{i+1}_{x1},{y1}_{x2},{y2}.jpg"
                    cv2.imwrite(img_path, crop_img)
                    img_paths.append(img_path)
            if whole_frame == 1:
                img_path = f"./{output_folder}/whole frames/{prefix}{species}_{timestamp}_{frame_number_start + frame_count}_{crop_counter}_whole.jpg"
                cv2.imwrite(img_path, frame)
            crop_counter += 1

        # If the random frame skip interval is activated add a random number to the counter or add the set frame skip interval
        if randomize == 1:
            if (frame_skip_loc - frame_count == 1):
                frame_count += 1
            else:
                frame_count += random.randint(1, max((frame_skip_loc - frame_count), 2))
        else:
            frame_count += frame_skip_loc

        # Read the next frame
        frame_to_read = frame_number_start + frame_count
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_to_read)
        success, frame = cap.read()

        # If the frame count is equal or larger than the amount of frames that comprises the duration of the visit end the loop
        if not (frame_count <= (visit_duration * fps)):

            # Release the video capture object and close all windows
            cap.release()
            cv2.destroyAllWindows()
            break

    # Return the resulting list of image paths for future reference
    return img_paths

async def yolo_preprocessing(img_paths, valid_annotations_array, index):
    model = YOLO('resources/yolo/best.pt')
    results = model(img_paths, save=False, imgsz=crop_size, conf=yolo_conf, save_txt=False, max_det=1, stream=True)
    for i, result in enumerate(results):
        boxes = result.boxes.data
        original_path = os.path.join(img_paths[i])
        create_dir(f"{output_folder}/empty")
        create_dir(f"{output_folder}/visitor")
        create_dir(f"{output_folder}/visitor/labels")
        empty_path = os.path.join(f"{output_folder}/empty", os.path.basename(img_paths[i]))
        visitor_path = os.path.join(f"{output_folder}/visitor", os.path.basename(img_paths[i]))
        label_path = os.path.join(f"{output_folder}/visitor/labels", os.path.basename(img_paths[i])[:-4])
        if len(result.boxes.data) > 0:
            shutil.move(original_path, visitor_path)
            with open(f"{label_path}.txt", 'w') as file:
                # Write the box_data to the file
                txt= []
                lst = result.boxes.xywhn[0].tolist()
                for item in lst:
                    txt_item = round(item, 6)
                    txt.append(txt_item)
                txt = str(txt)
                if any(len(row) < 7 for row in valid_annotations_array):
                    visitor_category = default_label_category
                else:
                    visitor_category = valid_annotations_array[index][6]
                file.write(f"{visitor_category} {txt.replace('[', '').replace(']', '').replace(',', '')}")
        else:
            shutil.move(original_path, empty_path)

def get_video_data(video_filepaths):
    global logger
    logger.debug(f"Running function get_video_data({video_filepaths})")
    # loop through time annotations and open corresponding video file
    # extract video data beforehand to save processing time
    video_data = []
    i: int
    for i, filepath in enumerate(video_filepaths):
        if filepath.endswith('.mp4'):
            video_start_time, video_end_time = get_video_start_end_times(video_filepaths[i])
            video_data_entry = [video_filepaths[i], video_start_time, video_end_time]
            video_data.append(video_data_entry)
    return video_data


# Define the mouse callback function to record the point of interest
def update_entries(index, original_points):
    global points_of_interest_entry
    global modified_frames
    global logger
    logger.debug(f"Running function update_entries({index}, {original_points})")
    for each in range(index + 1, len(points_of_interest_entry)):
        if len(points_of_interest_entry[each]) == 0:
            points_of_interest_entry[each] = points_of_interest_entry[index].copy()
        else:
            if points_of_interest_entry[each] == original_points:
                points_of_interest_entry[each] = points_of_interest_entry[index].copy()
            else:
                break
    first = 1
    for each in range(index, len(video_filepaths)):
        first = first - 1
        if first >= 0 and (
                index == 0 or not points_of_interest_entry[max(index - 1, 0)] == points_of_interest_entry[index]):
            modified_frames.append(each)
        update_button_image(frames[each].copy(), (max(each, 0) // 6), (each - ((max(each, 0) // 6) * 6)), first)

def get_mouse_position(event, x, y, flags, mode, i, j):
    global points_of_interest_entry
    global modified_frames
    if event == cv2.EVENT_LBUTTONUP:
        index = j + ((i) * 6)
        if flags & cv2.EVENT_FLAG_SHIFTKEY:
            closest_point = None
            closest_distance = float('inf')
            for point in points_of_interest_entry[index]:
                distance = math.sqrt((x - point[0]) ** 2 + (y - point[1]) ** 2)
                if distance < closest_distance:
                    closest_point = point
                    closest_distance = distance
            if closest_distance < 30:
                points_of_interest_entry[index].remove(closest_point)
            print(f"Flow: retrieved POIs: {points_of_interest_entry}")
            cv2.destroyAllWindows()
        elif flags & cv2.EVENT_FLAG_CTRLKEY and flags & cv2.EVENT_FLAG_ALTKEY:
            points_of_interest_entry[index] = []
            if index in modified_frames:
                modified_frames.remove(index)
            cv2.destroyAllWindows()
        else:
            if not mode == 0:
                points_of_interest_entry[index].append((x, y))
            else:
                points_of_interest_entry.append((x, y))
            print(f"Flow: retrieved POIs: {points_of_interest_entry}")
            cv2.destroyAllWindows()

def update_button_image(frame, i, j, first):
    global points_of_interest_entry
    global modified_frames
    global button_images
    global buttons
    global logger
    logger.debug(f"Running function update_button_image({i}, {j}, {first})")
    index = j + ((i) * 6)
    frame = frame.copy()

    # Draw roi area guides and offset overlap
    frame = draw_roi_offset_boundaries(frame, points_of_interest_entry[index], [-1,-1,-1,-1], True, False, True)

    if (first >= 0 or index in modified_frames) and (
            (index == 0 and not len(points_of_interest_entry[0]) == 0) or not points_of_interest_entry[
                                                                                  max(index - 1, 0)] ==
                                                                              points_of_interest_entry[index]):
        # Define the ROI
        frame = cv2.resize(frame, (276, 156), interpolation=cv2.INTER_AREA)
        height, width, channels = frame.shape
        x, y, w, h = 0, 0, height // 5, height // 5
        roi = frame[y:y + h, x:x + w]

        # Load the overlay image
        overlay = cv2.imread('resources/img/sp.png', cv2.IMREAD_UNCHANGED)

        # Resize the overlay image to fit the ROI
        overlay = cv2.resize(overlay, (w, h))

        # Merge the overlay image and the ROI
        alpha = overlay[:, :, 3] / 255.0
        alpha = cv2.merge([alpha, alpha, alpha])
        background = cv2.multiply(1.0 - alpha, roi, dtype=cv2.CV_8UC3)
        foreground = cv2.multiply(alpha, overlay[:, :, 0:3], dtype=cv2.CV_8UC3)
        result = cv2.add(background, foreground)

        # Copy the result back into the original image
        frame[y:y + h, x:x + w] = result
    pil_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(pil_img, mode='RGB')
    # Resize image to fit button
    pil_img = pil_img.resize((276, 156))
    # Convert PIL image to tkinter image
    img = ImageTk.PhotoImage(pil_img)
    button_images[index] = img
    buttons[i][j].configure(image=button_images[index])

def on_key_press(event):
    global b_l
    global b_r
    global u_l
    global u_r
    print("execd")
    if keyboard.is_pressed('left') and keyboard.is_pressed('down'):
        b_l = b_l * -1
    elif keyboard.is_pressed('right') and keyboard.is_pressed('down'):
        b_r = b_r * -1
    elif keyboard.is_pressed('left') and keyboard.is_pressed('up'):
        u_l = u_l * -1
    elif keyboard.is_pressed('right') and keyboard.is_pressed('up'):
        u_r = u_r * -1
    elif keyboard.is_pressed('q'):
        b_l = b_l * -1
        b_r = b_r * -1
        u_l = u_l * -1
        u_r = u_r * -1

def draw_roi_offset_boundaries(frame, list_of_ROIs, list_of_extremes_to_draw, draw_roi: bool, draw_extremes: bool, draw_overlap: bool):

    # Define variables
    u_l, u_r, b_l, b_r = list_of_extremes_to_draw
    rectangles = []
    labels = ['BR', 'UL', 'UR', 'BL']
    offsets = [(1, 1), (-1, -1), (1, -1), (-1, 1)]
    conditions = [b_r > 0, u_l > 0, u_r > 0, b_l > 0]
    pos_off = 0
    height, width, channels = frame.shape

    # For each point draw desired shapes
    for point in list_of_ROIs:

        # Draw basic roi area
        if draw_roi:
            x1, y1 = max(0, point[0] - (crop_size // 2)), max(0, point[1] - (crop_size // 2))
            x2, y2 = min(width, point[0] + (crop_size // 2)), min(height, point[1] + (crop_size // 2))
            cv2.rectangle(frame, (point[0] - 30, point[1] - 30), (point[0] + 30, point[1] + 30), (0, 255, 0), 2)
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 3)
            cv2.line(frame, (point[0], point[1]), (x1, y1), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            cv2.line(frame, (point[0], point[1]), (x1, y2), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            cv2.line(frame, (point[0], point[1]), (x2, y1), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)
            cv2.line(frame, (point[0], point[1]), (x2, y2), (0, 255, 255), thickness=2, lineType=cv2.LINE_AA)

        # Draw the offset extremes rectangles
        if draw_extremes or draw_overlap:
            for i, (label, (offset_x, offset_y), condition) in enumerate(zip(labels, offsets, conditions)):
                o_x1 = max(pos_off, min(((point[0] - crop_size // 2) + offset_x * offset_range) - pos_off,
                                        width - crop_size - pos_off))
                o_y1 = max(pos_off, min(((point[1] - crop_size // 2) + offset_y * offset_range) - pos_off,
                                        height - crop_size - pos_off))
                o_x2 = max(crop_size - pos_off,
                           min(((point[0] + crop_size // 2) + offset_x * offset_range) - pos_off,
                               width - pos_off))
                o_y2 = max(crop_size - pos_off,
                           min(((point[1] + crop_size // 2) + offset_y * offset_range) - pos_off,
                               height - pos_off))
                rectangles.append([(o_x1, o_y1), (o_x2, o_y2)])
                if condition:
                    cv2.rectangle(frame, (o_x1, o_y1), (o_x2, o_y2), (255, 229, 0), 2)

            # Draw overlap area
            if draw_overlap:

                # Find the overlapping area
                x_min = max(rect[0][0] for rect in rectangles)
                y_min = max(rect[0][1] for rect in rectangles)
                x_max = min(rect[1][0] for rect in rectangles)
                y_max = min(rect[1][1] for rect in rectangles)

                # Calculate the coordinates of the overlapping area
                cv2.rectangle(frame, (x_min, y_min), (x_max, y_max), (0, 255, 0), 2)
    return frame

def on_button_click(i, j, button_images):
    global points_of_interest_entry
    global logger
    global b_l
    global b_r
    global u_l
    global u_r
    logger.debug(f"Running function on_button_click({i}, {j})")
    index = j + ((i) * 6)
    mode = 1
    frame_tmp = frames[index].copy()
    b_l = -1
    b_r = -1
    u_l = -1
    u_r = -1
    # Ask the user if they want to select additional points of interest
    while True:
        original_points = points_of_interest_entry[index].copy()
        frame = frame_tmp.copy()
        list_of_extremes_to_draw = [u_l, u_r, b_l, b_r]

        # Draw a rectangle around the already selected points of interest
        frame = draw_roi_offset_boundaries(frame, points_of_interest_entry[index], list_of_extremes_to_draw, True, True, True)

        # Display the image with the rectangles marking the already selected points of interest
        cv2.imshow("Frame", frame)
        cv2.setWindowProperty("Frame", cv2.WND_PROP_TOPMOST, 1)
        screen_width, screen_height = cv2.getWindowImageRect("Frame")[2:]
        window_width, window_height = int(1 * screen_width), int(1 * screen_height)
        cv2.resizeWindow("Frame", window_width, window_height)
        cv2.moveWindow("Frame", int((screen_width // 2) - (window_width // 2)), 0)

        # Prompt the user to click on the next point of interest
        cv2.setMouseCallback("Frame",
                             lambda event, x, y, flags, mode: get_mouse_position(event, x, y, flags, mode, i, j), mode)

        # Wait for key press
        key = cv2.waitKey(0)

        # Define which keys will result in multiplying which variable by which value
        key_mappings = {
            27: ('esc', None),
            13: ('enter', None),
            ord('q'): ('u_l', -1),
            ord('w'): ('u_r', -1),
            ord('e'): ('b_l', -1),
            ord('r'): ('b_r', -1)
        }

        # If key is in the dictionary then check whether the value for this key is none - if so then it is esc or enter
        # and the window should be closed. If not then d othe multiplication.
        if key in key_mappings:
            action, value = key_mappings[key]
            if value is None:  # Check if the Esc key was pressed
                cv2.destroyAllWindows()
                break
            elif value is not None:
                globals()[action] *= value
        update_entries(index, original_points)

def load_video_frames():
    # Loop through each file in folder
    global frames
    global logger
    logger.debug(f"Running function load_video_frames()")
    frames = []
    if check_path(video_folder_path, 0):
        for filename in os.listdir(video_folder_path):
            if filename.endswith(".mp4"):  # Modify file extension as needed
                # Use OpenCV or other library to extract first frame of video
                # and add it to the frames list
                cap = cv2.VideoCapture(os.path.join(video_folder_path, filename))
                cap.set(cv2.CAP_PROP_POS_FRAMES, 25)
                ret, frame = cap.read()
                if ret:
                    frames.append(frame)
                else:
                    # If the read operation fails, add a default image
                    default_image = cv2.imread('resources/img/nf.png')
                    frames.append(default_image)
    else:
        # display message box with error message
        messagebox.showerror("Error", "Invalid video folder path")

def update_crop_mode(var):
    global crop_mode
    global logger
    logger.debug(f"Running function update_crop_mode({var})")
    crop_mode = var

def save_progress():
    global auto_processing
    global points_of_interest_entry
    global video_filepaths
    global logger
    logger.debug(f"Running function save_progress()")
    result = ask_yes_no("Do you want to save the settings? This will overwrite any previous saves.")
    if result:
        if check_path(video_folder_path, 0):
            # Create an in-memory file object
            filepath = os.path.join(video_folder_path, 'crop_information.pkl')
            with open(filepath, 'wb') as f:
                # Use the pickle module to write the data to the file
                pickle.dump([auto_processing.get(), points_of_interest_entry, video_filepaths], f)
        else:
            # display message box with error message
            messagebox.showerror("Error", "Invalid video folder path")

def load_progress():
    global auto_processing
    global points_of_interest_entry
    global video_filepaths
    global frames
    global loaded
    global auto
    global root
    global logger
    logger.debug(f"Running function load_progress()")
    result = ask_yes_no("Do you want to load settings? This will overwrite any unsaved progress.")
    if result:
        if check_path(video_folder_path, 0):
            # Create an in-memory file object
            filepath = os.path.join(video_folder_path, 'crop_information.pkl')
            if os.path.isfile(filepath):
                try:
                    if root.winfo_exists():
                        root.destroy()
                except:
                    print("Error: Unexpected, window destroyed before reference. Odpruženo.")
                points_of_interest_entry = []
                with open(filepath, 'rb') as f:
                    data = pickle.load(f)
                #print(data)
                auto = data[0]
                points_of_interest_entry = data[1].copy()
                video_filepaths_new = data[2].copy()
                set1 = set(video_filepaths_new)
                set2 = set(video_filepaths)
                if not set1 == set2:
                    messagebox.showinfo("Discrepancy detected",
                                        "The contents of the video folder have changed since the save has been made. Cannot load the progress. Please start over.")
                    reload_points_of_interest()
                else:
                    video_filepaths = []
                    video_filepaths = video_filepaths_new.copy()
                reload(1, False)
                # load_videos()
                # load_video_frames()
                # loaded = 1
                # open_ICCS_window()
            else:
                messagebox.showinfo("No save detected",
                                    "There are no save files in the current directory.")
        else:
            # display message box with error message
            messagebox.showerror("Error", "Invalid video folder path")


def open_ICCS_window():
    # Create tkinter window
    global ICCS_window
    global tree_allow
    global scaned_folders
    global video_folder_path
    global root
    global annotation_file_path
    global auto
    global logger
    logger.debug(f"Running function open_ICCS_window()")
    root = tk.Tk()
    ICCS_window = root
    root.focus()
    global gui_imgs
    gui_imgs = []

    def load_icon(path, size: tuple = (50, 50)):
        global gui_imgs
        # Load an image in the script
        img = (Image.open(path))

        # Resize the Image using resize method
        resized_image = img.resize(size)
        new_img = ImageTk.PhotoImage(resized_image)
        gui_imgs.append(new_img)
        return new_img

    j = 0
    root.title(f"Insect Communities Crop Suite - Folder: {os.path.basename(os.path.normpath(video_folder_path))} - Table: {os.path.basename(os.path.normpath(annotation_file_path))}")

    # Create frame for the rest
    outer_frame = tk.Frame(root)
    outer_frame.pack(side=tk.TOP, expand=True, fill=tk.BOTH)

    toolbar = tk.Frame(outer_frame)
    toolbar.pack(side=tk.TOP, fill=tk.BOTH)

    left_arrow = Image.open("resources/img/la.png")
    pil_img = left_arrow.resize((50, 50))
    pil_img.save("resources/img/la.png")
    left_arrow = ImageTk.PhotoImage(file="resources/img/la.png")
    if tree_allow and not scaned_folders.index(os.path.basename(os.path.normpath(video_folder_path))) == 0:
        btn_state = "normal"
    else:
        btn_state = "disabled"
    left_button = tk.Button(toolbar, image=load_icon("resources/img/la.png"), compound=tk.LEFT, text="Previous folder", padx=10, pady=5,
                            height=48, width=200, state=btn_state, command=lambda j=j: switch_folder("left"))
    left_button.image = load_icon("resources/img/la.png")
    left_button.grid(row=0, column=0, padx=0, pady=5, sticky="ew")

    menu_icon = Image.open("resources/img/mn.png")
    pil_img = menu_icon.resize((50, 50))
    pil_img.save("resources/img/mn.png")
    menu_icon = ImageTk.PhotoImage(file="resources/img/mn.png")

    menu_button = tk.Button(toolbar, image=menu_icon, compound=tk.LEFT, text="Menu", padx=10, pady=5, height=48,
                            command=lambda j=j: open_menu())
    menu_button.grid(row=0, column=1, padx=0, pady=5, sticky="ew")

    # frame for radio
    radio_frame = tk.Frame(toolbar)
    radio_frame.grid(row=0, column=2, padx=0, pady=5, sticky="ew")

    # create a tkinter variable to hold the selected value
    selected_option = tk.StringVar(value=crop_mode)

    one_icon = Image.open("resources/img/1.png")
    pil_img = one_icon.resize((50, 50))
    pil_img.save("resources/img/1.png")
    one_icon = ImageTk.PhotoImage(file="resources/img/1.png")

    two_icon = Image.open("resources/img/2.png")
    pil_img = two_icon.resize((50, 50))
    pil_img.save("resources/img/2.png")
    two_icon = ImageTk.PhotoImage(file="resources/img/2.png")

    three_icon = Image.open("resources/img/3.png")
    pil_img = three_icon.resize((50, 50))
    pil_img.save("resources/img/3.png")
    three_icon = ImageTk.PhotoImage(file="resources/img/3.png")

    # create the radio buttons and group them together
    rb1 = tk.Radiobutton(radio_frame, text="", image=one_icon, variable=selected_option, value=1, indicatoron=False,
                         height=56, width=116, font=("Arial", 17), command=lambda j_=j: update_crop_mode(1))
    rb2 = tk.Radiobutton(radio_frame, text="", image=two_icon, variable=selected_option, value=2, indicatoron=False,
                         height=56, width=116, font=("Arial", 17), command=lambda j=j: update_crop_mode(2))
    rb3 = tk.Radiobutton(radio_frame, text="", image=three_icon, variable=selected_option, value=3, indicatoron=False,
                         height=56, width=116, font=("Arial", 17), command=lambda j=j: update_crop_mode(3))

    # arrange the radio buttons in a horizontal layout using the grid geometry manager
    rb1.grid(row=0, column=0, sticky="ew")
    rb2.grid(row=0, column=1, sticky="ew")
    rb3.grid(row=0, column=2, sticky="ew")
    radio_frame.grid_columnconfigure(0, weight=1, minsize=50)
    radio_frame.grid_columnconfigure(1, weight=1, minsize=50)
    radio_frame.grid_columnconfigure(2, weight=1, minsize=50)

    on_image = tk.PhotoImage(width=116, height=57)
    off_image = tk.PhotoImage(width=116, height=57)
    on_image.put(("green",), to=(0, 0, 56, 56))
    off_image.put(("red",), to=(57, 0, 115, 56))
    global auto_processing
    auto_processing = tk.IntVar(value=0)
    auto_processing.set(0)
    cb1 = tk.Checkbutton(toolbar, image=off_image, selectimage=on_image, indicatoron=False, onvalue=1, offvalue=0,
                         variable=auto_processing)
    cb1.grid(row=0, column=3, padx=0, pady=5, sticky="ew")

    gears_icon = Image.open("resources/img/au.png")
    pil_img = gears_icon.resize((50, 50))
    pil_img.save("resources/img/au.png")
    gears_icon = ImageTk.PhotoImage(file="resources/img/au.png")

    auto_button = tk.Button(toolbar, image=gears_icon, compound=tk.LEFT, text="Automatic evaluation", padx=10,
                            pady=5, height=48, command=lambda j=j: auto_processing.set(1 - auto_processing.get()))
    auto_button.grid(row=0, column=4, padx=0, pady=5, sticky="ew")

    fl_icon = Image.open("resources/img/fl.png")
    pil_img = fl_icon.resize((50, 50))
    pil_img.save("resources/img/fl.png")
    fl_icon = ImageTk.PhotoImage(file="resources/img/fl.png")

    fl_button = tk.Button(toolbar, image=fl_icon, compound=tk.LEFT, text="Select video folder", padx=10, pady=5,
                          height=48, command=lambda j=j: change_video_folder())
    fl_button.grid(row=0, column=5, padx=0, pady=5, sticky="ew")

    et_icon = Image.open("resources/img/et.png")
    pil_img = et_icon.resize((50, 50))
    pil_img.save("resources/img/et.png")
    et_icon = ImageTk.PhotoImage(file="resources/img/et.png")

    et_button = tk.Button(toolbar, image=et_icon, compound=tk.LEFT, text="Select Excel table", padx=10, pady=5,
                          height=48, command=lambda j=j: change_excel_path())
    et_button.grid(row=0, column=6, padx=0, pady=5, sticky="ew")

    ocr_icon = Image.open("resources/img/ocr.png")
    pil_img = ocr_icon.resize((50, 50))
    pil_img.save("resources/img/ocr.png")
    ocr_icon = ImageTk.PhotoImage(file="resources/img/ocr.png")

    ocr_button = tk.Button(toolbar, image=ocr_icon, compound=tk.LEFT, text="OCR", padx=10, pady=5, height=48, width=100, command=lambda j=j: set_ocr_roi(video_filepaths[0]))
    ocr_button.grid(row=0, column=7, padx=0, pady=5, sticky="ew")

    right_arrow = Image.open("resources/img/ra.png")
    pil_img = right_arrow.resize((50, 50))
    pil_img.save("resources/img/ra.png")
    right_arrow = ImageTk.PhotoImage(file="resources/img/ra.png")
    if tree_allow and not (scaned_folders.index(os.path.basename(os.path.normpath(video_folder_path))) + 1) == len(
            scaned_folders):
        btn_state1 = "normal"
    else:
        btn_state1 = "disabled"
    right_button = tk.Button(toolbar, image=right_arrow, compound=tk.RIGHT, text="Next folder", padx=10, pady=5,
                             height=48, width=200, state=btn_state1, command=lambda j=j: switch_folder("right"))
    right_button.grid(row=0, column=8, padx=0, pady=5, sticky="ew")

    toolbar.grid_columnconfigure(0, weight=2, minsize=50)
    toolbar.grid_columnconfigure(1, weight=3, minsize=50)
    toolbar.grid_columnconfigure(2, weight=4, minsize=150)
    toolbar.grid_columnconfigure(3, weight=4, minsize=50)
    toolbar.grid_columnconfigure(4, weight=1, minsize=50)
    toolbar.grid_columnconfigure(5, weight=4, minsize=50)
    toolbar.grid_columnconfigure(6, weight=4, minsize=50)
    toolbar.grid_columnconfigure(7, weight=4, minsize=50)
    toolbar.grid_columnconfigure(8, weight=2, minsize=50)

    # Create a canvas to hold the buttons
    canvas = tk.Canvas(outer_frame)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Create a scrollbar for the canvas
    scrollbar = tk.Scrollbar(outer_frame, orient=tk.VERTICAL, command=canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Configure the canvas to use the scrollbar
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Create target frame for the rest
    target_frame = tk.Frame(outer_frame)
    target_frame.pack(side=tk.TOP)
    toolbar.lift(target_frame)

    global button_images
    global buttons
    button_images = []
    buttons = []
    rows = math.ceil(len(video_filepaths) / 6)
    #per_row = math.ceil(len(video_filepaths) // rows)
    label_frame = tk.Frame(target_frame, width=50)
    label_frame.pack(side=tk.LEFT)
    for i in range(rows):
        hour_1st = ((i) * 6)
        if (((i) * 6) + 6) + 1 <= len(video_filepaths):
            hour_2nd = max(((i) * 6) + 6, len(video_filepaths) - 1)
        else:
            hour_2nd = (len(video_filepaths) % 6) - 1
        text_label = tk.Label(label_frame,
                              text=f"{os.path.basename(video_filepaths[hour_1st])[-9:-7]}-{os.path.basename(video_filepaths[hour_2nd])[-9:-7]}",
                              font=("Arial", 15), background=root.cget('bg'))
        text_label.pack(side=tk.TOP, padx=30, pady=67)
    for i in range(rows):
        # Loop through first 24 frames and create buttons with images
        button_frame = tk.Frame(target_frame)
        button_frame.pack(side=tk.TOP)
        row = []
        for j in range(6):
            if (j + ((i) * 6)) < len(video_filepaths):
                # Convert frame to PIL image
                pil_img = cv2.cvtColor(frames[j + ((i) * 6)], cv2.COLOR_BGR2RGB)
                pil_img = Image.fromarray(pil_img, mode='RGB')
                # Resize image to fit button
                pil_img = pil_img.resize((276, 156))
                # Convert PIL image to tkinter image
                tk_img = ImageTk.PhotoImage(pil_img)
                button_images.append(tk_img)
                # Create button with image and add to button frame
                button = tk.Button(button_frame, image=button_images[j + ((i) * 6)],
                                   command=lambda i=i, j=j: on_button_click(i, j, button_images))
                button.grid(row=i, column=j, sticky="w")
                row.append(button)
            else:
                # Create a dummy frame to fill in the grid
                new_img = Image.new("RGBA", (276, 156), (0, 0, 0, 0))
                new_img = ImageTk.PhotoImage(new_img)
                dummy_frame = tk.Button(button_frame, image=new_img, foreground="white", state="disabled")
                dummy_frame.grid(row=i, column=j, sticky='w')
                row.append(dummy_frame)
        buttons.append(row)

    # Update the canvas to show the buttons
    canvas.create_window((0, 0), window=target_frame, anchor=tk.NW)
    target_frame.update_idletasks()

    # Configure the canvas to show the entire frame
    canvas.configure(scrollregion=canvas.bbox("all"))

    # Bind mouse wheel event to canvas
    canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(-1 * (event.delta // 120), "units"))

    bottom_left_panel = tk.Frame(root, pady=0)
    bottom_left_panel.pack(side=tk.LEFT)

    fl_button = tk.Button(bottom_left_panel, image=fl_icon, compound=tk.LEFT, text="", padx=10, pady=5, width=60, height=58,
                          command=lambda j=j: os.startfile(video_folder_path))
    fl_button.pack(side=tk.LEFT)

    bottom_right_panel = tk.Frame(root, pady=0)
    bottom_right_panel.pack(side=tk.RIGHT)

    et_button = tk.Button(bottom_right_panel, image=et_icon, compound=tk.LEFT, text="", padx=10, pady=5, width=60, height=58,
                          command=lambda j=j: os.startfile(annotation_file_path))
    et_button.pack(side=tk.RIGHT)


    ef_icon = Image.open("resources/img/ef.png")
    pil_img = ef_icon.resize((50, 50))
    pil_img.save("resources/img/ef.png")
    ef_icon = ImageTk.PhotoImage(file="resources/img/ef.png")
    ef_button = tk.Button(bottom_right_panel, image=ef_icon, compound=tk.LEFT, text="", padx=0, pady=5, width=60, height=58,
                          command=lambda j=j: os.startfile(os.path.dirname(annotation_file_path)))
    ef_button.pack(side=tk.RIGHT)

    bottom_toolbar = tk.Frame(root, pady=5)
    bottom_toolbar.pack(side=tk.TOP, fill=tk.BOTH)

    save_icon = Image.open("resources/img/sv.png")
    pil_img = save_icon.resize((50, 50))
    pil_img.save("resources/img/sv_1.png")
    save_icon = ImageTk.PhotoImage(file="resources/img/sv_1.png")
    # create a Button widget with the save icon as its image
    save_button = tk.Button(bottom_toolbar, text="Save", image=save_icon, compound=tk.LEFT, padx=10, pady=5, width=300,
                            height=48, command=lambda j=j: save_progress())

    crop_icon = Image.open("resources/img/cr.png")
    pil_img = crop_icon.resize((50, 50))
    pil_img.save("resources/img/cr_1.png")
    crop_icon = ImageTk.PhotoImage(file="resources/img/cr_1.png")
    # create a Button widget with the save icon as its image
    crop_button = tk.Button(bottom_toolbar, text="Crop", image=crop_icon, compound=tk.LEFT, padx=10, pady=5, width=300,
                            height=48, command=lambda j=j: crop_engine())

    sort_icon = Image.open("resources/img/so.png")
    pil_img = sort_icon.resize((50, 50))
    pil_img.save("resources/img/so.png")
    sort_icon = ImageTk.PhotoImage(file="resources/img/so.png")
    # create a Button widget with the save icon as its image
    sort_button = tk.Button(bottom_toolbar, text="Sort", image=sort_icon, compound=tk.LEFT, padx=10, pady=5, width=300,
                            height=48, command=lambda j=j: sort_engine())

    load_icon = Image.open("resources/img/lo.png")
    pil_img = load_icon.resize((50, 50))
    pil_img.save("resources/img/lo.png")
    load_icon = ImageTk.PhotoImage(file="resources/img/lo.png")
    # create a Button widget with the save icon as its image
    load_button = tk.Button(bottom_toolbar, text="Load", image=load_icon, compound=tk.LEFT, padx=10, pady=5, width=300,
                            height=48, command=lambda j=j: load_progress())
    # Specify the column for each button
    save_button.grid(row=0, column=1, sticky="ew")
    crop_button.grid(row=0, column=2, sticky="ew")
    sort_button.grid(row=0, column=3, sticky="ew")
    load_button.grid(row=0, column=4, sticky="ew")

    # Add padding between the buttons
    bottom_toolbar.grid_columnconfigure(0, weight=1)
    bottom_toolbar.grid_columnconfigure(1, weight=2)
    bottom_toolbar.grid_columnconfigure(2, weight=2)
    bottom_toolbar.grid_columnconfigure(3, weight=2)
    bottom_toolbar.grid_columnconfigure(4, weight=2)
    bottom_toolbar.grid_columnconfigure(5, weight=1)

    # Get the screen width and height
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    root_width = root.winfo_width()

    # update
    global loaded
    if loaded == 1:
        for each in range(len(video_filepaths)):
            first = -1
            update_button_image(frames[each].copy(), (max(each, 0) // 6), (each - ((max(each, 0) // 6) * 6)), first)
        # if auto is defined, set the auto processing to that value otherwise set it to 0
        if auto is None:
            auto_processing.set(0)
        else:
            auto_processing.set(auto)

    # Set the window size to fit the entire screen
    root.geometry(f"{screen_width}x{screen_height-70}+-10+0")
    #root.state('zoomed')
    root.mainloop()


def initialise():
    global loaded
    global tree_allow
    global modified_frames
    global video_folder_path
    global annotation_file_path
    global auto
    global logger
    log_write()
    logger.debug("Running function initialise()")
    config_read()
    pytesseract.pytesseract.tesseract_cmd = ocr_tesseract_path
    auto = 0
    loaded = 0
    tree_allow = False
    modified_frames = []
    video_folder_path, annotation_file_path = scan_default_folders(scan_folders, crop_mode)
    while not check_path(video_folder_path, 0):
        video_folder_path, scaned_folders, tree_allow = get_video_folder(video_folder_path, 1)
    annotation_file_path = get_excel_path(annotation_file_path, 1, video_folder_path, crop_mode)
    load_videos()
    create_dir(output_folder)
    create_dir(f"./{output_folder}/whole frames/")
    reload_points_of_interest()
    load_video_frames()

def filter_array_by_visitors(valid_annotations_array):
    global filtered_array
    # filter window
    filter_window = tk.Tk()
    filter_window.title("Filter Visitors")
    filter_window.wm_attributes("-topmost", 1)

    def apply_filter():
        global filtered_array
        results = []
        print(checkbox_vars)
        for i, var in enumerate(checkbox_vars):
            checkbox_value = var.get()
            print(checkbox_value)
            checkbox_text = checkboxes[i].cget("text")
            print(checkbox_text)
            dropdown_value = selected_items[i].get()
            print(dropdown_value)
            if not dropdown_value[0].isdigit():
                dropdown_value = "6. empty"
            results.append([int(checkbox_value), checkbox_text, int(dropdown_value[0])])
        results = [row for row in results if row[0] != 0]
        if yolo_processing == 1:
            for row in results:
                if row[2] == 6:
                    filter_window.quit()
                    filter_window.destroy()
                    messagebox.showinfo("Warning",
                                        f"One of the selected groups of visitors was not assigned a correct category for labelling. Please try again.")
                    filtered_array = []
                    return
        print(results)
        selected_values = [value.get() for value in checkbox_vars]
        filtered_array = []
        for row in valid_annotations_array:
            if row[5] in [result[1] for result in results]:
                matching_result = next(result for result in results if result[1] == row[5])
                filtered_row = row + [matching_result[2]]
                filtered_array.append(filtered_row)
        print(filtered_array)  # Do something with the filtered array
        filter_window.quit()
        filter_window.destroy()

    def handle_selection(index, selection):
        if selection in allowed_items:
            selected_items[index].set(selection)
            label_values[index].config(text=selection)
        else:
            selected_items[index].set("")
            label_values[index].config(text="")

    # Get unique values from column index 5
    column_5_values = [row[5] for row in valid_annotations_array]
    unique_values = set(column_5_values)
    allowed_items = ["0. Hymenoptera", "1. Diptera", "2. Lepidoptera", "3. Coleoptera", "4. Other"]
    if len(unique_values) > 0:
        checkbox_vars = []
        checkboxes = []
        selected_items = []
        label_values = []
        for i, value in enumerate(unique_values):
            var = tk.StringVar(filter_window, value=0)  # Pass 'filter_window' as the 'master' argument
            checkbox_vars.append(var)

            checkbox = tk.Checkbutton(filter_window, text=value, variable=var)
            checkbox.grid(row=i, column=0, sticky='w')
            checkboxes.append(checkbox)

            selected_item = tk.StringVar(filter_window)
            selected_item.set("------------")
            selected_items.append(selected_item)

            option_menu = tk.OptionMenu(filter_window, selected_item, *allowed_items,
                                        command=lambda selection, index=i: handle_selection(index, selection))
            option_menu.config(width=35)
            option_menu.grid(row=i, column=1, sticky='nsew')

            label_value = tk.Label(filter_window, text="", width=40)
            label_value.grid(row=i, column=2, sticky='nsew')
            label_values.append(label_value)

        apply_button = tk.Button(filter_window, text="Apply Filter", command=apply_filter)
        apply_button.grid(row=len(unique_values), column=0, columnspan=3, sticky='nsew', pady=(30, 0))

        #Set properties and start window
        filter_window.update()
        screen_width = filter_window.winfo_screenwidth()
        screen_height = filter_window.winfo_screenheight()
        window_width = filter_window.winfo_reqwidth()
        window_height = filter_window.winfo_reqheight()
        x_pos = int((screen_width - window_width) / 2)
        y_pos = int((screen_height - window_height) / 2)
        filter_window.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")
        filter_window.mainloop()
    else:
        print("Flow: No visitor items to filter by. Processing all visits.")
        filtered_array = []
        for row in valid_annotations_array:
                filtered_row = row + [0]
                filtered_array.append(filtered_row)
        print(filtered_array)  # Do something with the filtered array
        filter_window.quit()
        filter_window.destroy()

def crop_engine():
    global points_of_interest_entry
    global video_filepaths
    global cap
    global fps
    global frame_number_start
    global visit_duration
    global crop_mode
    global root
    global loaded
    global logger
    global whole_frame
    global visit_index
    logger.debug("Running function crop_engine()")
    result = ask_yes_no("Do you want to start the cropping process?")
    if result:
        if len(points_of_interest_entry[0]) == 0:
            messagebox.showinfo("Warning", "No points of interest selected. Please select at least one POI.")
            reload(1, False)
            return
        valid_annotations_array = []
        valid_annotation_data_entry = []
        print(f"Flow: Start cropping on the following videos: {video_filepaths}")
        root.withdraw()
        if not crop_mode == 3:
            video_ok = check_path(video_folder_path, 0)
            excel_ok = check_path(annotation_file_path, 1)
            if not video_ok or not excel_ok:
                messagebox.showinfo("Warning",
                                    f"Unspecified path to a video folder or a valid Excel file.")
                reload(1, False)
                return
            if crop_mode == 1:
                video_data = get_video_data(video_filepaths)
                annotation_data_array, visitor_id = load_excel_table(annotation_file_path)
            if crop_mode == 2:
                annotation_data_array = load_csv(annotation_file_path)
            if annotation_data_array is None:
                messagebox.showinfo("Warning", f"Attempted to fix errors in the selected Excel file. Attempt failed. Please fix the errors manually and try again.")
                reload(1, False)
                return
            if crop_mode == 2:
                video_filepaths_temp = video_filepaths
                video_filepaths = []
                for index, list in enumerate(annotation_data_array):
                    for filepath in video_filepaths_temp:
                        filename = os.path.basename(filepath)  # get just the filename from the full path
                        if annotation_data_array[index][1][:-9] in filename:
                            if datetime.timedelta() <= datetime.datetime.strptime(
                                    annotation_data_array[index][1][-8:-3], '%H_%M') - datetime.datetime.strptime(
                                    filename[-9:-4], '%H_%M') <= datetime.timedelta(minutes=15):
                                video_filepaths.append(filepath)
                video_data = get_video_data(video_filepaths)
            print(f"Flow: Start cropping according to the following annotations: {annotation_data_array}")
            print(f"Flow: Start cropping with the following video data: {video_data}")
            for index, list in enumerate(annotation_data_array):
                print(' '.join(["Flow: Annotation number:", str(index + 1)]))
                annotation_time = pd.to_datetime(annotation_data_array[index][1], format='%Y%m%d_%H_%M_%S')
                for i, list in enumerate(video_data):
                    print(f"{video_data[i][1]} <= {annotation_time} <= {video_data[i][2]}")
                    if video_data[i][1] <= annotation_time <= video_data[i][2]:
                        for each in range(len(annotation_data_array[index])):
                            valid_annotation_data_entry.append(annotation_data_array[index][each])
                        for each in range(3):
                            valid_annotation_data_entry.append(video_data[i][each])
                        if crop_mode == 1:
                            valid_annotation_data_entry.append(visitor_id[index][0])
                        print(f"Flow: Relevant annotations: {valid_annotation_data_entry}")
                        valid_annotations_array.append(valid_annotation_data_entry)
                        valid_annotation_data_entry = []
            if filter_visitors == 1:
                filter_array_by_visitors(valid_annotations_array)
                if len(filtered_array) > 0:
                    valid_annotations_array = filtered_array
                else:
                    print("No visitors of the selected type found.")
                    reload(1, False)
                    return
            for index in range(len(valid_annotations_array)):
                print(f"Processing item {index}")
                visit_index = index
                annotation_time = pd.to_datetime(valid_annotations_array[index][1], format='%Y%m%d_%H_%M_%S')
                annotation_offset = (annotation_time - valid_annotations_array[index][3]).total_seconds()
                cap = cv2.VideoCapture(valid_annotations_array[index][2])
                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                fps = cap.get(cv2.CAP_PROP_FPS)
                visit_duration = (min(((annotation_offset * fps) + (int(valid_annotations_array[index][0]) * fps)),
                                      total_frames) - (annotation_offset * fps)) // fps
                frame_number_start = int(annotation_offset * fps)
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number_start)
                success, frame = cap.read()
                img_paths = asyncio.run(generate_frames(frame, success, os.path.basename(valid_annotations_array[index][2]),
                                    video_filepaths.index(valid_annotations_array[index][2])))
                #loop.close()
                #img_paths = generate_frames(frame, success, os.path.basename(valid_annotations_array[index][2]),
                               # video_filepaths.index(valid_annotations_array[index][2]))
                if yolo_processing == 1:
                    asyncio.run(yolo_preprocessing(img_paths, valid_annotations_array, index))
        else:
            video_ok = check_path(video_folder_path, 0)
            if not video_ok:
                messagebox.showinfo("Warning",
                                    f"Unspecified path to a video folder.")
                reload(1, False)
                return
            orig_wf = whole_frame
            whole_frame = 1
            for i, filepath in enumerate(video_filepaths):
                cap = cv2.VideoCapture(video_filepaths[i])
                fps = cap.get(cv2.CAP_PROP_FPS)
                visit_duration = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) // fps
                frame_number_start = 2
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number_start)
                success, frame = cap.read()
                img_paths = generate_frames(frame, success, os.path.basename(video_filepaths[i]), i)
            whole_frame = orig_wf
    reload(1, False)
    # try:
    #     if root.winfo_exists():
    #         root.destroy()
    # except:
    #     print("Error: Unexpected, window destroyed before reference. Odpruženo.")
    # loaded = 1
    # open_ICCS_window()

def sort_engine():
    global logger
    logger.debug("Running function sort_engine()")
    # Ask user if they want to run sorting script
    run_sorting = ask_yes_no("Do you want to Running function the sorting script on the generated images?")
    if run_sorting:
        sort_script_path = "sort.py"
        if os.path.exists(sort_script_path):
            # subprocess.run(['python', f'{sort_script_path}'])
            subprocess.call([sys.executable, f'{sort_script_path}', "--subprocess"])
        else:
            print("Error: sorting script not found.")


def open_menu():
    global output_folder
    global scan_folders
    global crop_mode
    global frame_skip
    global frames_per_visit
    global filter_visitors
    global yolo_processing
    global default_label_category
    global yolo_conf
    global randomize
    global whole_frame
    global cropped_frames
    global crop_size
    global offset_range
    global prefix
    global end_values
    global logger
    logger.debug("Running function open_menu()")
    # Create the Tkinter window
    window = tk.Tk()
    window.title("Menu")
    window.wm_attributes("-topmost", 1)
    # Create the labels and input fields
    label_text = ["Output folder path:", "Scan default folders:", "Filename prefix:", "Default crop mode:",
                  "Frames to skip:", "Frames per visit:", "Filter visitors:", "Yolo processing", "Default label category", "Yolo conf. tresh.", "Randomize interval:", "Export whole frames:", "Export cropped frames:",
                  "Crop size:", "Offset size:"]
    labels = []
    fields = []
    outer_frame = tk.Frame(window, pady=20)
    outer_frame.pack(side=tk.TOP, fill=tk.BOTH)
    for i in range(15):
        label = tk.Label(outer_frame, text=f"{label_text[i]}")
        label.grid(row=i, column=0, padx=10)
        labels.append(label)

        field = tk.Entry(outer_frame, width=120)
        field.grid(row=i, column=1, padx=10)
        fields.append(field)

    # Create the save button
    def save_fields():
        global output_folder
        global scan_folders
        global crop_mode
        global frame_skip
        global frames_per_visit
        global filter_visitors
        global yolo_processing
        global default_label_category
        global yolo_conf
        global randomize
        global whole_frame
        global cropped_frames
        global crop_size
        global offset_range
        global prefix
        global end_values
        end_values = []
        variable_mappings = [
            ('output_folder', str),
            ('scan_folders', str),
            ('prefix', str),
            ('crop_mode', int),
            ('frame_skip', int),
            ('frames_per_visit', int),
            ('filter_visitors', int),
            ('yolo_processing', int),
            ('default_label_category', int),
            ('yolo_conf', float),
            ('randomize', int),
            ('whole_frame', int),
            ('cropped_frames', int),
            ('crop_size', int),
            ('offset_range', int)
        ]

        end_values = [var_type(fields[i].get()) for i, (var_name, var_type) in enumerate(variable_mappings)]
        print(end_values)
        for i, (var_name, _) in enumerate(variable_mappings):
            globals()[var_name] = end_values[i]
        config_write()
        create_dir(output_folder)
        create_dir(f"./{output_folder}/whole frames/")
        window.destroy()

    save_button = tk.Button(outer_frame, text="Save", command=save_fields)
    save_button.grid(row=16, column=0, columnspan=2)

    # Set initial values for the input fields
    initial_values = [output_folder, scan_folders, prefix, crop_mode, frame_skip, frames_per_visit, filter_visitors, yolo_processing, default_label_category, yolo_conf, randomize, whole_frame,
                      cropped_frames, crop_size, offset_range]
    for i in range(15):
        fields[i].insert(0, str(initial_values[i]))

    # Start the Tkinter event
    window.update()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    window_width = window.winfo_reqwidth()
    window_height = window.winfo_reqheight()
    x_pos = int((screen_width - window_width) / 2)
    y_pos = int((screen_height - window_height) / 2)
    window.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")
    window.mainloop()


def config_write():
    global ocr_tesseract_path
    global video_folder_path
    global annotation_file_path
    global output_folder
    global scan_folders
    global crop_mode
    global frame_skip
    global frames_per_visit
    global filter_visitors
    global yolo_processing
    global default_label_category
    global yolo_conf
    global randomize
    global whole_frame
    global cropped_frames
    global crop_size
    global offset_range
    global config
    global prefix
    global end_values
    global logger
    logger.debug("Running function config_write()")
    config = configparser.ConfigParser()
    config.read('settings_crop.ini')

    # Update values in the config file
    #config.set('Resource Paths', 'OCR_tesseract_path', ocr_tesseract_path)
    #config.set('Resource Paths', 'video_folder_path', video_folder_path)
    #config.set('Resource Paths', 'annotation_file_path', annotation_file_path)
    config.set('Resource Paths', 'output_folder', output_folder)
    config.set('Workflow settings', 'Scan_default_folders', scan_folders)
    config.set('Crop settings', 'crop_mode', str(crop_mode))
    config.set('Crop settings', 'crop_interval_frames', str(frame_skip))
    config.set('Crop settings', 'frames_per_visit', str(frames_per_visit))
    config.set('Crop settings', 'filter_visitors', str(filter_visitors))
    config.set('Crop settings', 'yolo_processing', str(yolo_processing))
    config.set('Crop settings', 'default_label_category', str(default_label_category))
    config.set('Crop settings', 'yolo_conf', str(yolo_conf))
    config.set('Crop settings', 'randomize_interval', str(randomize))
    config.set('Crop settings', 'export_whole_frame', str(whole_frame))
    config.set('Crop settings', 'export_crops', str(cropped_frames))
    config.set('Crop settings', 'crop_size', str(crop_size))
    config.set('Crop settings', 'random_offset_range', str(offset_range))
    config.set('Crop settings', 'filename_prefix', str(prefix))

    # Save changes to the config file
    with open('settings_crop.ini', 'w', encoding='utf-8') as configfile:
        config.write(configfile)

def log_write():
    global logger
    # Create a logger instance
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    # Create a file handler that logs all messages, and set its formatter
    file_handler = logging.FileHandler('runtime.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # Create a console handler that logs only messages with level INFO or higher, and set its formatter
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(levelname)s: %(message)s')
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)

    # Write log messages
    # logger.debug('This message will be written to the file only.')
    # logger.info('This message will be written to both the file and the console.')
    # logger.warning('This message will be written to both the file and the console.')
    # logger.error('This message will be written to both the file and the console.')
    # logger.critical('This message will be written to both the file and the console.')

def reload(is_loaded, reload_POIs):
    global loaded
    global root
    global logger
    logger.debug(f"Running function reload({is_loaded}, {reload_POIs})")
    load_videos()
    if reload_POIs:
        reload_points_of_interest()
    load_video_frames()
    try:
        if root.winfo_exists():
            root.destroy()
    except:
        print("Error: Unexpected, window destroyed before reference. Odpruženo.")
    loaded = is_loaded
    open_ICCS_window()

# Main body of the script

# Profile the generate_frames function
profiler = cProfile.Profile()
profiler.enable()

initialise()
open_ICCS_window()

# Stop profiling
profiler.disable()
profiler.print_stats()
